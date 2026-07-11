#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试用例记录表 — TestCase Viewer & Recorder
============================================
将 Excel 测试用例逐条展示为清晰的 HTML 卡片页面，
执行后记录测试结果和备注，自动回写 Excel。
支持搜索筛选 + 汇总统计页。

技术栈：Python Flask + openpyxl + 浏览器前端
适配：Mac / Windows
"""

import sys
import os
import json
import webbrowser
import threading
import datetime
import re
from pathlib import Path
from copy import copy

# 阶段2：Repository 抽象层（API 通过此接口访问数据，不直碰 STATE/Excel）
from app.services.repository import create_repository

# ============================================================
# 0. 依赖自检 & 自动安装
# ============================================================
def _ensure_deps():
    missing = []
    for pkg in ['flask', 'openpyxl']:
        try:
            __import__(pkg if pkg != 'flask' else 'flask')
        except ImportError:
            missing.append(pkg)

    if missing:
        import subprocess
        print(f"⏳ 首次运行，正在安装必要组件: {', '.join(missing)} ...")
        print("   （仅需一次，请稍候）\n")
        try:
            subprocess.check_call(
                [sys.executable, '-m', 'pip', 'install', '--quiet'] + missing,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            print("✅ 组件安装完成！\n")
        except Exception as e:
            print(f"❌ 自动安装失败: {e}")
            print("   请手动运行: pip install flask openpyxl")
            input("\n按回车键退出...")
            sys.exit(1)

_ensure_deps()

from flask import Flask, request, jsonify, render_template_string, render_template
import openpyxl

# ============================================================
# 1. 配置常量
# ============================================================
BASE_DIR = Path(__file__).parent.resolve()
TESTCASES_DIR = BASE_DIR / 'testcases'
PORT = 8765
MEMORY_FILE = TESTCASES_DIR / '_memory.json'  # 记录原始文件名

# 列名智能识别规则（支持中英文常见表头）
COLUMN_PATTERNS = {
    'id':           ['用例编号', '编号', '用例ID', '序号', 'No.', '编号ID', '用例号', 'NO'],
    'title':        ['用例标题', '用例名称', '标题', '测试点', '名称', '测试项', '测试标题', '功能点', '测试内容'],
    'precondition': ['前置条件', '预置条件', '前提条件', '准备条件', '预设', '环境准备'],
    'steps':        ['测试步骤', '步骤', '操作步骤', '执行步骤', '测试过程', '操作过程', '测试操作'],
    'expected':     ['预期结果', '期望结果', '预计结果', '预期输出', '期望输出'],
    'purpose':      ['测试目的', '目的', '测试目标', '测试说明', '测试意图'],
    'priority':     ['优先级', '重要程度', '严重程度', 'P级', '重要级别'],
    'module':       ['所属模块', '模块', '功能模块', '测试模块', '需求模块', '系统模块'],
    'category':     ['用例类型', '测试类型', '测试分类', '分类'],
    'result_col':   ['测试结果', '执行结果', 'Pass/Fail', 'pass/fail', 'PASS/FAIL',
                     'Result', 'Status', '结果', '状态', '测试状态', '用例状态',
                     'Pass', 'Fail', '通过', '不通过', '失败', '阻塞', '跳过'],
    'remark_col':   ['实际结果', '备注', '测试备注', 'Remark', '实际现象',
                     '现象', 'Actual Result', '测试结果备注', '测试备注'],
}

# 保存时写入的列名（如 Excel 中没有则自动新建）
RESULT_COL = '测试结果'
REMARK_COL = '测试现象备注'
TIME_COL   = '执行时间'


# ============================================================
# 2. Excel 操作
# ============================================================


def find_excel_file():
    """在 testcases 目录中查找 Excel 文件。
    优先读 _memory.json 中记录的文件，其次找第一个 Excel。"""
    if not TESTCASES_DIR.exists():
        TESTCASES_DIR.mkdir(parents=True)
        return None

    # 收集所有 Excel 文件
    excels = []
    for pattern in ['*.xlsx', '*.xls']:
        for f in TESTCASES_DIR.glob(pattern):
            if not f.name.startswith('~$') and not f.name.startswith('.'):
                excels.append(f)

    if not excels:
        return None

    # 优先用记忆中的文件
    remembered = read_memory()
    for f in excels:
        if f.name == remembered:
            return f

    # 降级：返回第一个
    return excels[0]


def read_memory():
    """读取记忆文件，返回原始文件名"""
    if MEMORY_FILE.exists():
        try:
            import json as _json
            with open(MEMORY_FILE, 'r', encoding='utf-8') as f:
                data = _json.load(f)
                return data.get('original_name', '')
        except Exception:
            pass
    return ''


def write_memory(original_name):
    """写入记忆文件，记录原始文件名"""
    import json as _json
    MEMORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(MEMORY_FILE, 'w', encoding='utf-8') as f:
        _json.dump({'original_name': original_name}, f, ensure_ascii=False)


def validate_excel(filepath):
    """校验 Excel 格式是否合法（有表头，有数据）"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        has_usable_sheet = False
        all_empty = True
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            if any(
                cell is not None and str(cell).strip()
                for row in rows
                for cell in row
            ):
                all_empty = False
            if len(rows) >= 2:
                header_idx, headers = _find_header_row(ws)
                has_header = bool(headers) and any(str(h).strip() for h in headers)
                has_data = _count_data_rows(ws, header_idx) > 0 if header_idx else False
                if has_header and has_data:
                    has_usable_sheet = True
                    break
        wb.close()
        if all_empty:
            return False, 'Excel 文件为空，请上传包含表头和数据的测试用例文件'
        if not has_usable_sheet:
            return False, 'Excel 至少需要一个包含表头和数据的 Sheet'
        return True, ''
    except Exception as e:
        return False, f'无法读取 Excel 文件，文件可能已损坏: {str(e)}'


def has_test_results():
    """检查当前活跃文件是否已有测试结果（用于覆盖前确认）。

    文件来源优先级：STATE['filepath'] → find_excel_file() → None。
    不再依赖已废弃的 _active.xlsx 固定文件名。
    """
    filepath = STATE.get('filepath')
    if not filepath:
        found = find_excel_file()
        if not found:
            return False
        filepath = str(found)
    if not Path(filepath).exists():
        return False
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = [str(ws.cell(row=1, column=c + 1).value).strip()
                   if ws.cell(row=1, column=c + 1).value is not None else ''
                   for c in range(ws.max_column)]
        result_col_idx = None
        for i, h in enumerate(headers):
            if h == RESULT_COL:
                result_col_idx = i + 1
                break
        if result_col_idx is None:
            wb.close()
            return False
        # 检查是否有任何一行已有结果
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=result_col_idx).value
            if val and str(val).strip():
                wb.close()
                return True
        wb.close()
        return False
    except Exception:
        return False


def detect_columns(headers):
    """将 Excel 表头智能映射到标准字段。
    优先精确匹配，避免模糊匹配的交叉误匹配。
    每个字段只匹配第一个命中的列（避免后面的列覆盖前面的）。"""
    mapping = {}
    matched_fields = set()
    for i, h in enumerate(headers):
        h_str = str(h).strip() if h else ''
        if h_str == '':
            continue

        # 第一遍：只做精确匹配
        matched = False
        for field, patterns in COLUMN_PATTERNS.items():
            if field in matched_fields:
                continue
            if h_str in patterns:
                mapping[field] = i
                matched_fields.add(field)
                matched = True
                break
        if matched:
            continue

        # 第二遍：模糊匹配（表头包含模式关键词 或 模式词包含表头）
        # 但仅在关键词长度 >= 3 时才模糊匹配，避免 "ID" 等短词误匹配
        for field, patterns in COLUMN_PATTERNS.items():
            if field in matched_fields:
                continue
            for pat in patterns:
                if len(pat) < 3:
                    continue  # 太短的关键词不做模糊匹配
                if pat in h_str or h_str in pat:
                    mapping[field] = i
                    matched_fields.add(field)
                    matched = True
                    break
            if matched:
                break

    mapping['_headers'] = headers
    return mapping


def _find_col_index(headers, column_names):
    """在 header list 中查找列名，支持单个名称或别名列表。
    返回 0-based index，找不到返回 None"""
    if isinstance(column_names, str):
        column_names = [column_names]
    for i, h in enumerate(headers):
        if h in column_names:
            return i
    return None


def _find_header_row_for_file(filepath):
    """为文件智能定位表头行（复用 _find_header_row 逻辑）"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    result = _find_header_row(ws)
    wb.close()
    return result


def _detect_header_row_index(filepath):
    """返回智能定位到的表头行号（1-based）"""
    idx, _ = _find_header_row_for_file(filepath)
    return idx if idx else 1


def _find_header_row(ws):
    """智能定位真正的表头行。
    自动跳过顶部的大标题、说明文本、空行等非数据行。
    返回 (表头行号, 表头内容列表)，表头行号为 1-based。"""
    raw_rows = list(ws.iter_rows(values_only=True, max_row=min(ws.max_row, 100)))
    if not raw_rows:
        return None, []

    def _clean(v):
        return str(v).strip() if v is not None else ''

    # 策略：找第一行"像表头"的行 —— 该行有效非空单元格 >= 2
    # 且该行下一行有数据（不是空行）
    for i, row in enumerate(raw_rows):
        filled = sum(1 for c in row if _clean(c))
        if filled >= 2:
            # 检查下一行是否有数据
            if i + 1 < len(raw_rows):
                next_filled = sum(1 for c in raw_rows[i+1] if _clean(c))
                if next_filled >= 1:
                    return (i + 1, [_clean(c) for c in row])
            else:
                # 最后一行就当它是表头好了
                return (i + 1, [_clean(c) for c in row])
    # 兜底：用第一行
    return (1, [_clean(c) for c in raw_rows[0]])


def _count_data_rows(ws, header_idx):
    """从表头行下一行开始，统计有效数据行数。"""
    count = 0
    for r in range(header_idx + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c + 1).value for c in range(ws.max_column)]
        if any(v is not None and str(v).strip() for v in row_vals):
            count += 1
    return count


CORE_FIELDS = {'id', 'title', 'steps', 'expected', 'precondition',
               'purpose', 'priority', 'module', 'category', 'result_col', 'remark_col'}
NAME_BONUS_KEYWORDS = ['用例', 'case', 'testcase', 'test case', 'test-case', 'cases', 'testcases']


def _classify_sheet(ws):
    """判定单个 Sheet 是测试用例 Sheet 还是参考信息 Sheet。"""
    raw_rows = list(ws.iter_rows(values_only=True, max_row=min(ws.max_row, 100)))
    header_idx, headers = _find_header_row(ws)
    mapping = detect_columns(headers)
    core_count = len([f for f in mapping if f in CORE_FIELDS])

    name_bonus = 0
    sname_lower = ws.title.lower()
    for kw in NAME_BONUS_KEYWORDS:
        if kw in sname_lower:
            name_bonus = 10
            break

    min_core = 2 if name_bonus > 0 else 3
    data_rows = _count_data_rows(ws, header_idx) if header_idx else 0
    sheet_type = 'testcase' if len(raw_rows) >= 2 and data_rows > 0 and core_count >= min_core else 'note'

    return {
        'name': ws.title,
        'sheet_name': ws.title,
        'sheet_type': sheet_type,
        'header_row': header_idx or 1,
        'headers': headers,
        'mapping': mapping,
        'case_count': data_rows if sheet_type == 'testcase' else 0,
        'core_field_count': core_count,
    }


def _parse_testcase_sheet(ws, meta):
    """把一个已分类为测试用例的 Sheet 解析成用例 dict 列表。"""
    all_rows = list(ws.iter_rows(values_only=True))
    header_idx = meta['header_row']
    headers = meta['headers']
    mapping = meta['mapping']
    data_rows = all_rows[header_idx:]

    result_col_idx = _find_col_index(headers, COLUMN_PATTERNS.get('result_col', [SAVE_COLUMNS['result']]))
    remark_col_idx = _find_col_index(headers, COLUMN_PATTERNS.get('remark_col', [SAVE_COLUMNS['actual_result']]))

    testcases = []
    for row_idx_in_data, row in enumerate(data_rows):
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        excel_row_number = header_idx + 1 + row_idx_in_data
        tc = {
            '_row': excel_row_number,
            '_sheet_name': meta['name'],
            '_header_row': header_idx,
            '_headers': headers,
            '_mapping': mapping,
        }

        for i, value in enumerate(row):
            tc[f'col_{i}'] = str(value).strip() if value is not None else ''
            tc[f'_header_{i}'] = headers[i] if i < len(headers) else f'列{i+1}'

        for field, col_idx in mapping.items():
            if field != '_headers' and col_idx < len(row):
                raw_val = row[col_idx]
                tc[field] = str(raw_val).strip() if raw_val is not None else ''

        if result_col_idx is not None and result_col_idx < len(row):
            tc['_saved_result'] = str(row[result_col_idx]).strip() if row[result_col_idx] is not None else ''
        else:
            tc['_saved_result'] = ''

        if remark_col_idx is not None and remark_col_idx < len(row):
            tc['_saved_actual_result'] = str(row[remark_col_idx]).strip() if row[remark_col_idx] is not None else ''
        else:
            tc['_saved_actual_result'] = ''

        for extra_key, col_name in [('_saved_tester', '测试人员'),
                                    ('_saved_bug_id', 'BugID'),
                                    ('_saved_bug_frequency', 'Bug频率'),
                                    ('_saved_issue_time', '问题时间')]:
            extra_col = None
            for i, h in enumerate(headers):
                if h == col_name:
                    extra_col = i
                    break
            if extra_col is not None and extra_col < len(row):
                tc[extra_key] = str(row[extra_col]).strip() if row[extra_col] is not None else ''
            else:
                tc[extra_key] = ''

        search_parts = []
        for i in range(len(row)):
            cell_val = tc.get(f'col_{i}', '')
            if cell_val and cell_val != 'None':
                search_parts.append(cell_val)
        tc['_search_text'] = ' '.join(search_parts)
        testcases.append(tc)

    return testcases


def read_all_sheets(filepath):
    """读取 Excel 全部 Sheet，分类为用例 Sheet 与参考信息 Sheet。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = []
    testcase_sheets = []
    note_sheets = []
    testcases = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        meta = _classify_sheet(ws)
        if meta['sheet_type'] == 'testcase':
            sheet_cases = _parse_testcase_sheet(ws, meta)
            meta['case_count'] = len(sheet_cases)
            testcases.extend(sheet_cases)
            testcase_sheets.append({
                'name': meta['name'],
                'case_count': meta['case_count'],
                'header_row': meta['header_row'],
            })
        else:
            note_sheets.append({
                'name': meta['name'],
                'header_row': meta['header_row'],
            })
        sheets.append(meta)

    first_testcase_sheet = next((s for s in sheets if s['sheet_type'] == 'testcase'), None)
    mapping = first_testcase_sheet['mapping'] if first_testcase_sheet else {}
    headers = first_testcase_sheet['headers'] if first_testcase_sheet else []
    sheet_name = first_testcase_sheet['name'] if first_testcase_sheet else (wb.sheetnames[0] if wb.sheetnames else '')
    header_row = first_testcase_sheet['header_row'] if first_testcase_sheet else 1
    wb.close()

    return {
        'testcases': testcases,
        'mapping': mapping,
        'headers': headers,
        'sheet_name': sheet_name,
        'header_row': header_row,
        'sheets': sheets,
        'testcase_sheets': testcase_sheets,
        'note_sheets': note_sheets,
    }


def _pick_best_sheet(wb):
    """在多个 Sheet 中自动选择真正的测试用例 Sheet。
    判断标准：
      1. Sheet 名称包含"用例/Case/TestCase"等关键词的优先
      2. 必须具备 ≥ 2 个核心用例字段（id/steps/expected/precondition/priority 等）
      3. 优先选已含「测试结果」列的 Sheet（说明用户已在填写）
      4. 同条件下选核心字段数量最多的那个
      5. 再同条件下选数据行最多的那个
    不满足条件 2 的 Sheet 视为补充说明，不会被选中。
    """
    best_sheet = wb.active
    best_score = (-1, -1, False, 0)   # (name_bonus, core_field_count, has_result, data_rows)

    for sname in wb.sheetnames:
        ws = wb[sname]
        raw_rows = list(ws.iter_rows(values_only=True, max_row=min(ws.max_row, 100)))
        if len(raw_rows) < 2:
            continue

        # 定位表头行
        header_idx = None
        for i, row in enumerate(raw_rows):
            filled = sum(1 for c in row if c is not None and str(c).strip())
            if filled >= 2:
                if i + 1 < len(raw_rows):
                    next_filled = sum(1 for c in raw_rows[i+1] if c is not None and str(c).strip())
                    if next_filled >= 1:
                        header_idx = i + 1
                        break
                else:
                    header_idx = i + 1
                    break
        if header_idx is None:
            continue

        header_row_vals = [str(c).strip() if c is not None else '' for c in raw_rows[header_idx - 1]]

        # 通过 detect_columns 识别有几个核心字段
        mapping = detect_columns(header_row_vals)
        core_count = len([f for f in mapping if f in CORE_FIELDS])

        # 名称匹配加分：名称包含"用例"等关键词的 sheet 给最高优先级
        sname_lower = sname.lower()
        name_bonus = 0
        for kw in NAME_BONUS_KEYWORDS:
            if kw in sname_lower:
                name_bonus = 10  # 最高优先级加分
                break

        # 对名称匹配"用例"的 sheet，放宽核心字段要求（≥2 即可）
        # 对名称不匹配"用例"的 sheet，仍需 ≥3 个核心字段
        min_core = 2 if name_bonus > 0 else 3
        if core_count < min_core:
            continue

        cnt = _count_data_rows(ws, header_idx)
        has_result = (SAVE_COLUMNS['result'] in header_row_vals)

        # 比较：名称匹配加分 > 核心字段数 > 有无结果列 > 数据行数
        score = (name_bonus, core_count, has_result, cnt)
        if score > best_score:
            best_score = score
            best_sheet = ws

    return best_sheet


def read_testcases(filepath):
    """读取 Excel，返回（用例列表, 列映射, 表头列表, sheet名称, 表头行号）。
    智能选择最佳 Sheet，自动跳过首行标题、说明文本等非表头行。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 智能选择最佳 Sheet
    ws = _pick_best_sheet(wb)
    sheet_name = ws.title

    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        wb.close()
        return [], {}, [], sheet_name, 1

    # 智能定位表头行
    header_idx, headers = _find_header_row(ws)
    # 数据行从表头的正下方一行开始
    # header_idx 是 1-based 行号，all_rows 是 0-based
    # all_rows[header_idx] 即表头行的下一行（第一条数据）
    data_start = header_idx  # 0-based index，第一条数据行在 all_rows 中的位置
    data_rows = all_rows[data_start:]

    mapping = detect_columns(headers)

    # 找到结果和备注列（使用别名列表，兼容 Pass/Fail 等常见表头）
    result_col_idx = _find_col_index(headers, COLUMN_PATTERNS.get('result_col', [SAVE_COLUMNS['result']]))
    remark_col_idx = _find_col_index(headers, COLUMN_PATTERNS.get('remark_col', [SAVE_COLUMNS['actual_result']]))

    # 解析数据行（从表头的下一行开始）
    testcases = []
    for row_idx_in_data, row in enumerate(data_rows):
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        # Excel 实际行号（1-based）
        # header_idx = 表头行号 (1-based)
        # row_idx_in_data = data_rows 中的索引 (0 = 第一条数据)
        excel_row_number = header_idx + 1 + row_idx_in_data
        tc = {'_row': excel_row_number}

        for i, value in enumerate(row):
            tc[f'col_{i}'] = str(value).strip() if value is not None else ''
            tc[f'_header_{i}'] = headers[i] if i < len(headers) else f'列{i+1}'

        for field, col_idx in mapping.items():
            if field != '_headers' and col_idx < len(row):
                raw_val = row[col_idx]
                tc[field] = str(raw_val).strip() if raw_val is not None else ''

        # 读入已保存的测试结果（从 result_col 列读取）
        if result_col_idx is not None and result_col_idx < len(row):
            tc['_saved_result'] = str(row[result_col_idx]).strip() if row[result_col_idx] is not None else ''
        else:
            tc['_saved_result'] = ''

        # 读入已保存的实际结果（从 remark_col 列读取——即 Excel 中的"实际结果"列）
        if remark_col_idx is not None and remark_col_idx < len(row):
            tc['_saved_actual_result'] = str(row[remark_col_idx]).strip() if row[remark_col_idx] is not None else ''
        else:
            tc['_saved_actual_result'] = ''

        # 读入已保存的 BugID、Bug频率、问题时间、测试人员
        # 这些字段通过列名直接从 header 查找（与 result_col 相同的查找方式）
        for extra_key, col_name in [('_saved_tester', '测试人员'),
                                      ('_saved_bug_id', 'BugID'),
                                      ('_saved_bug_frequency', 'Bug频率'),
                                      ('_saved_issue_time', '问题时间')]:
            extra_col = None
            for i, h in enumerate(headers):
                if h == col_name:
                    extra_col = i
                    break
            if extra_col is not None and extra_col < len(row):
                tc[extra_key] = str(row[extra_col]).strip() if row[extra_col] is not None else ''
            else:
                tc[extra_key] = ''

        # 构建搜索文本（所有字段拼接，用于全文搜索）
        search_parts = []
        for i in range(len(row)):
            field_key = f'col_{i}'
            cell_val = tc.get(field_key, '')
            if cell_val and cell_val != 'None':
                search_parts.append(cell_val)
        tc['_search_text'] = ' '.join(search_parts)

        testcases.append(tc)

    wb.close()
    return testcases, mapping, headers, sheet_name, header_idx


# 保存时写入的列名（如 Excel 中没有则自动新建）
SAVE_COLUMNS = {
    'result':        '测试结果',
    'actual_result': '实际结果',
    'tester':        '测试人员',
    'bug_id':        'BugID',
    'bug_frequency': 'Bug频率',
    'issue_time':    '问题时间',
    'exec_time':     '执行时间',
}


def save_result(filepath, row_number, header_row_idx, sheet_name, result, actual_result,
                tester, bug_id, bug_frequency, issue_time):
    """将测试结果写入 Excel 指定行和指定 Sheet。
    优先复用已有的结果列（支持 Pass/Fail 等别名），找不到时才新建列。
    新建列时自动复制原表头的字体、背景、边框、对齐样式，保持表格风格统一。
    header_row_idx: 智能定位到的表头行号（1-based）
    sheet_name: 目标 Sheet 名称"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]  # 使用指定的 Sheet，而不是 active

    # 使用智能定位到的表头行
    # 实际保存列名：测试结果 / 实际结果 / 测试人员 / BugID / Bug频率 / 问题时间 / 执行时间
    max_col = max(ws.max_column, 1)
    header_cells = [ws.cell(row=header_row_idx, column=c + 1) for c in range(max_col)]
    headers_row = [str(c.value).strip() if c.value is not None else '' for c in header_cells]

    # 读取所有行以便追加时也能正确
    all_rows_vals = list(ws.iter_rows(min_row=header_row_idx, values_only=True))
    # 确保 headers_row 至少覆盖到最宽的行
    for r in all_rows_vals:
        while len(headers_row) < len(r):
            headers_row.append('')

    # 取第一个已有表头单元格作为样式模板
    style_source_cell = None
    for c in header_cells:
        if c.value is not None:
            style_source_cell = c
            break

    def _copy_style(target_cell, source_cell):
        """将 source_cell 的字体、填充、边框、对齐复制到 target_cell"""
        if source_cell is None:
            return
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.number_format:
            target_cell.number_format = copy(source_cell.number_format)

    def find_or_create_col(col_name, aliases=None):
        """查找列，支持主名称+别名。未找到则在最右侧新建，并复制表头样式。"""
        search_names = [col_name]
        if aliases:
            search_names = list(aliases) + search_names
        for i, h in enumerate(headers_row):
            if h in search_names:
                return i + 1
        # 没找到：新建列
        new_col = len(headers_row) + 1
        new_header_cell = ws.cell(row=header_row_idx, column=new_col, value=col_name)
        headers_row.append(col_name)
        # 复制表头样式
        _copy_style(new_header_cell, style_source_cell)
        # 同时复制一个数据行模板样式（取原表头正下方第一行数据格）
        data_source_cell = ws.cell(row=header_row_idx + 1, column=1)
        if data_source_cell and data_source_cell.value is not None:
            _copy_style(new_header_cell, data_source_cell)
        return new_col

    # 按需查找或创建列（结果列和备注列使用别名匹配已有列）
    result_col = find_or_create_col(SAVE_COLUMNS['result'], aliases=COLUMN_PATTERNS.get('result_col'))
    actual_col = find_or_create_col(SAVE_COLUMNS['actual_result'], aliases=COLUMN_PATTERNS.get('remark_col'))
    tester_col = find_or_create_col(SAVE_COLUMNS['tester'])
    bug_id_col = find_or_create_col(SAVE_COLUMNS['bug_id'])
    bug_freq_col = find_or_create_col(SAVE_COLUMNS['bug_frequency'])
    issue_time_col = find_or_create_col(SAVE_COLUMNS['issue_time'])
    exec_time_col = find_or_create_col(SAVE_COLUMNS['exec_time'])

    now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 写入数据，同时复制数据行模板样式
    data_source_cell = ws.cell(row=header_row_idx + 1, column=1)
    cells_to_write = [
        (row_number, result_col, result),
        (row_number, actual_col, actual_result),
        (row_number, tester_col, tester),
        (row_number, bug_id_col, bug_id),
        (row_number, bug_freq_col, bug_frequency),
        (row_number, issue_time_col, issue_time),
        (row_number, exec_time_col, now_str),
    ]
    for r, c, val in cells_to_write:
        cell = ws.cell(row=r, column=c, value=val)
        _copy_style(cell, data_source_cell)

    wb.save(filepath)
    wb.close()


def build_display_fields(tc, mapping, headers):
    """构建前端展示用的字段列表（有序）。
    所有 Excel 列都展示，即使值为空也不省略。"""
    fields = []

    ordered_specials = [
        ('title',        '\U0001f4c4 用例标题', 'field-section field-title'),
        ('id',           '\U0001f4cb 用例编号', 'field-id'),
        ('priority',     '⚡ 优先级',   'field-meta'),
        ('module',       '\U0001f4c1 所属模块', 'field-meta'),
        ('category',     '\U0001f5c2 用例分类', 'field-meta'),
        ('precondition', '\U0001f527 前置条件', 'field-section field-section-pre'),
        ('steps',        '\U0001f4dd 测试步骤', 'field-section field-steps'),
        ('expected',     '✅ 预期结果', 'field-section field-expected'),
        ('purpose',      '\U0001f4a1 测试目的', 'field-section field-purpose'),
    ]

    displayed_cols = set()

    for field, label, css_class in ordered_specials:
        if field in mapping:
            col_idx = mapping[field]
            val = tc.get(field, '')
            fields.append({
                'label': label,
                'value': val if val else '',
                'css_class': css_class,
            })
            displayed_cols.add(col_idx)

    # 其余未展示的列也全部输出（包括空值）
    # 注意：不跳过任何原始 Excel 列，包括原始就存在的 Pass/Fail、备注等列
    for i, h in enumerate(headers):
        if i in displayed_cols:
            continue
        val = tc.get(f'col_{i}', '')
        if val == 'None':
            val = ''
        fields.append({
            'label': f'\U0001f4c4 {h}',
            'value': val if val else '',
            'css_class': 'field-section',
        })

    return fields


# ============================================================
# 3. Flask 应用 & API
# ============================================================
app = Flask(__name__, template_folder="app/templates", static_folder="app/static")

STATE = {
    'testcases': [],
    'mapping': {},
    'headers': [],
    'sheets': [],
    'testcase_sheets': [],
    'note_sheets': [],
    'filepath': None,
    'filename': '',
    'sheet_name': None,      # 实际使用的 sheet 名称
    'header_row': 1,         # 智能定位到的表头行号
}

# 阶段2 T6/T8：Repository 实例（config 驱动，换数据库只改这里）
repo = create_repository(os.environ.get('STORAGE_BACKEND', 'excel'))


def _apply_loaded_data(filepath, loaded):
    """把 read_all_sheets 的结果写入全局 STATE，保留旧字段兼容。"""
    STATE['testcases'] = loaded['testcases']
    STATE['mapping'] = loaded['mapping']
    STATE['headers'] = loaded['headers']
    STATE['sheets'] = loaded['sheets']
    STATE['testcase_sheets'] = loaded['testcase_sheets']
    STATE['note_sheets'] = loaded['note_sheets']
    STATE['filepath'] = str(filepath)
    STATE['filename'] = Path(filepath).name
    STATE['sheet_name'] = loaded['sheet_name']
    STATE['header_row'] = loaded['header_row']


@app.route('/')
def index():
    return render_template("index.html")


@app.route('/api/init')
def api_init():
    display_name = read_memory() or STATE.get('filename', '')
    return jsonify({
        'loaded': len(STATE['testcases']) > 0,
        'filename': display_name,
        'total': len(STATE['testcases']),
        'has_active': find_excel_file() is not None,
        'sheet_count': len(STATE.get('sheets', [])),
        'testcase_sheet_count': len(STATE.get('testcase_sheets', [])),
        'note_sheet_count': len(STATE.get('note_sheets', [])),
    })


@app.route('/api/sheets')
def api_sheets():
    """返回当前 Excel 文件中所有 Sheet 的分类信息。"""
    filepath = STATE.get('filepath') or find_excel_file()
    if filepath is None:
        return jsonify({'error': '未找到活跃文件，请先上传测试用例 Excel'}), 404

    try:
        if not STATE.get('sheets') or str(filepath) != STATE.get('filepath'):
            loaded = read_all_sheets(filepath)
            _apply_loaded_data(filepath, loaded)

        return jsonify({
            'testcase_sheets': STATE.get('testcase_sheets', []),
            'note_sheets': STATE.get('note_sheets', []),
            'total_sheets': len(STATE.get('sheets', [])),
        })
    except Exception as e:
        return jsonify({'error': f'读取 Sheet 信息失败: {str(e)}'}), 500


@app.route('/api/titles')
def api_titles():
    """返回全部用例的索引、编号、标题列表，供前端下拉框快速跳转使用。"""
    tcs = STATE['testcases']
    result = []
    for i, tc in enumerate(tcs):
        title = tc.get('title', '') or tc.get('name', '') or tc.get('col_2', '') or '(无标题)'
        id_val = tc.get('id', '') or tc.get('col_0', '') or f'#{i + 1}'
        result.append({
            'index': i,
            'id': str(id_val).strip(),
            'title': str(title).strip(),
        })
    return jsonify(result)

@app.route('/api/testcase/<int:index>')
def api_testcase(index):
    if index < 0 or index >= len(STATE['testcases']):
        return jsonify({'error': '索引超出范围'}), 404

    tc = STATE['testcases'][index].copy()
    tc['_index'] = index
    tc['_display_fields'] = build_display_fields(
        tc,
        tc.get('_mapping', STATE['mapping']),
        tc.get('_headers', STATE['headers']),
    )
    tc['_total'] = len(STATE['testcases'])

    return jsonify(tc)


@app.route('/api/all-status')
def api_all_status():
    """获取所有用例的执行状态"""
    if not STATE['filepath'] or not STATE['testcases']:
        return jsonify([])

    try:
        wb = openpyxl.load_workbook(STATE['filepath'], data_only=True)
        statuses = []
        for tc in STATE['testcases']:
            sn = tc.get('_sheet_name') or STATE.get('sheet_name')
            ws = wb[sn] if sn else wb.active
            header_row = tc.get('_header_row') or STATE.get('header_row', 1)
            headers_row = [
                str(ws.cell(row=header_row, column=c + 1).value).strip()
                if ws.cell(row=header_row, column=c + 1).value is not None else ''
                for c in range(ws.max_column)
            ]
            result_col = None
            for i, h in enumerate(headers_row):
                if h == RESULT_COL:
                    result_col = i + 1
                    break
            row = tc['_row']
            if result_col:
                cell_val = ws.cell(row=row, column=result_col).value
                statuses.append(str(cell_val).strip() if cell_val else '')
            else:
                statuses.append('')

        wb.close()
        return jsonify(statuses)
    except Exception:
        return jsonify([''] * len(STATE['testcases']))


@app.route('/api/search', methods=['POST'])
def api_search():
    """搜索筛选 API

    请求体示例:
    {
        "keyword": "登录",           // 全文搜索关键词（可选）
        "result_filter": "",         // 按结果筛选: ""全部 / "通过" / "失败" / "阻塞" / "跳过" / "未执行"
        "priority_filter": "",       // 按优先级筛选（可选）
        "module_filter": "",         // 按模块筛选（可选）
        "page": 0,                   // 分页页码（0-based）
        "page_size": 50              // 每页条数
    }

    返回:
    {
        "results": [ { index, id, title, module, priority, result, remark, _search_snippet } ],
        "total_matched": 150,
        "page": 0,
        "page_size": 50
    }
    """
    data = request.get_json() or {}
    keyword = (data.get('keyword') or '').strip().lower()
    result_filter = (data.get('result_filter') or '').strip()
    priority_filter = (data.get('priority_filter') or '').strip()
    module_filter = (data.get('module_filter') or '').strip()
    page = max(0, int(data.get('page', 0)))
    page_size = min(200, max(1, int(data.get('page_size', 50))))

    # 收集所有用例的状态（从内存中）
    matched = []
    for i, tc in enumerate(STATE['testcases']):
        # 全文搜索
        if keyword:
            if keyword not in tc.get('_search_text', '').lower():
                continue

        # 结果筛选
        if result_filter:
            saved = tc.get('_saved_result', '')
            if result_filter == '未执行':
                if saved != '' and saved != 'None':
                    continue
            elif saved != result_filter:
                continue

        # 优先级筛选
        if priority_filter:
            tc_priority = tc.get('priority', '')
            if priority_filter not in tc_priority:
                continue

        # 模块筛选
        if module_filter:
            tc_module = tc.get('module', '')
            if module_filter not in tc_module:
                continue

        # 构建搜索摘要（关键词高亮上下文）
        snippet = ''
        if keyword:
            search_text = tc.get('_search_text', '')
            idx = search_text.lower().find(keyword)
            if idx >= 0:
                start = max(0, idx - 30)
                end = min(len(search_text), idx + len(keyword) + 50)
                snippet = search_text[start:end]
                if start > 0:
                    snippet = '...' + snippet
                if end < len(search_text):
                    snippet = snippet + '...'

        matched.append({
            'index': i,
            'id': tc.get('id', ''),
            'title': tc.get('title', ''),
            'name': tc.get('col_1', '') or tc.get('title', '') or tc.get('col_0', ''),  # 用例名称（优先从货架名称/标题取）
            'module': tc.get('module', ''),
            'priority': tc.get('priority', ''),
            'result': tc.get('_saved_result', ''),
            'remark': tc.get('_saved_remark', ''),
            'snippet': snippet,
        })

    total_matched = len(matched)
    start = page * page_size
    end = start + page_size
    page_results = matched[start:end]

    return jsonify({
        'results': page_results,
        'total_matched': total_matched,
        'page': page,
        'page_size': page_size,
        'total_pages': max(1, (total_matched + page_size - 1) // page_size),
    })


@app.route('/api/summary')
def api_summary():
    """汇总统计 API

    返回:
    {
        "total": 200,
        "counts": { "通过": 120, "失败": 15, "阻塞": 5, "跳过": 10, "未执行": 50 },
        "by_priority": { "P0": { "通过":5, "失败":2, "阻塞":0, "跳过":0, "未执行":3 }, ... },
        "by_module": { "登录模块": { "通过":10, "失败":1, ... }, ... },
        "execution_rate": 75.0,       // 已执行百分比
        "pass_rate": 80.0,            // 通过率（通过的/已执行）
    }
    """
    tcs = STATE['testcases']

    counts = {'通过': 0, '失败': 0, '阻塞': 0, '跳过': 0, '未执行': 0}
    by_priority = {}
    by_module = {}

    for tc in tcs:
        result = tc.get('_saved_result', '')
        if not result or result == 'None':
            result = '未执行'

        # 全局计数
        if result in counts:
            counts[result] += 1
        else:
            counts['未执行'] += 1

        # 按优先级分组
        pri = tc.get('priority', '未标注')
        if not pri or pri == 'None':
            pri = '未标注'
        if pri not in by_priority:
            by_priority[pri] = {'通过': 0, '失败': 0, '阻塞': 0, '跳过': 0, '未执行': 0}
        by_priority[pri][result] += 1

        # 按模块分组
        mod = tc.get('module', '未分类')
        if not mod or mod == 'None':
            mod = '未分类'
        if mod not in by_module:
            by_module[mod] = {'通过': 0, '失败': 0, '阻塞': 0, '跳过': 0, '未执行': 0}
        by_module[mod][result] += 1

    total = len(tcs)
    executed = total - counts['未执行']
    execution_rate = round((executed / total) * 100, 1) if total > 0 else 0
    pass_rate = round((counts['通过'] / executed) * 100, 1) if executed > 0 else 0

    return jsonify({
        'total': total,
        'counts': counts,
        'by_priority': by_priority,
        'by_module': by_module,
        'execution_rate': execution_rate,
        'pass_rate': pass_rate,
    })


@app.route('/api/filter-options')
def api_filter_options():
    """获取所有可用的筛选选项（优先级列表、模块列表）"""
    priorities = set()
    modules = set()

    for tc in STATE['testcases']:
        p = tc.get('priority', '')
        if p and p != 'None':
            priorities.add(p)
        m = tc.get('module', '')
        if m and m != 'None':
            modules.add(m)

    return jsonify({
        'priorities': sorted(priorities),
        'modules': sorted(modules),
    })


@app.route('/api/save', methods=['POST'])
def api_save():
    """保存测试结果"""
    data = request.get_json()
    index = data.get('index')
    result = data.get('result', '')
    actual_result = data.get('actual_result', '')
    tester = data.get('tester', '')
    bug_id = data.get('bug_id', '')
    bug_frequency = data.get('bug_frequency', '')
    issue_time = data.get('issue_time', '')

    if index is None or index < 0 or index >= len(STATE['testcases']):
        return jsonify({'success': False, 'error': '索引无效'}), 400
    if not STATE['filepath']:
        return jsonify({'success': False, 'error': '未找到 Excel 文件'}), 400

    try:
        tc = STATE['testcases'][index]
        row_number = tc['_row']
        header_row_idx = tc.get('_header_row') or STATE.get('header_row', 1)
        sheet_name = tc.get('_sheet_name') or STATE.get('sheet_name')
        # 阶段2 T6：通过 Repository 保存（不直接调 save_result）
        repo.save(
            filepath=STATE['filepath'],
            row_number=row_number,
            header_row_idx=header_row_idx,
            sheet_name=sheet_name,
            result=result,
            actual_result=actual_result,
            tester=tester,
            bug_id=bug_id,
            bug_frequency=bug_frequency,
            issue_time=issue_time,
        )

        # 更新内存
        STATE['testcases'][index]['_saved_result'] = result
        STATE['testcases'][index]['_saved_actual_result'] = actual_result
        STATE['testcases'][index]['_saved_tester'] = tester
        STATE['testcases'][index]['_saved_bug_id'] = bug_id
        STATE['testcases'][index]['_saved_bug_frequency'] = bug_frequency
        STATE['testcases'][index]['_saved_issue_time'] = issue_time

        return jsonify({'success': True})
    except PermissionError:
        return jsonify({
            'success': False,
            'error': '无法写入 Excel 文件。请关闭 Excel 程序后重试。'
        }), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reload')
def api_reload():
    """重新加载 Excel 文件，pick 外部修改（新增/删除用例等）"""
    global STATE
    filepath = find_excel_file()
    if filepath is None:
        return jsonify({
            'success': False,
            'error': '未找到活跃文件，请先上传测试用例 Excel'
        }), 404

    try:
        loaded = read_all_sheets(filepath)
        _apply_loaded_data(filepath, loaded)
        display_name = read_memory() or filepath.name

        return jsonify({
            'success': True,
            'filename': display_name,
            'total': len(STATE['testcases']),
            'sheet_count': len(STATE.get('sheets', [])),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/upload', methods=['POST'])
def api_upload():
    """上传 Excel 文件，保留原始文件名并写入记忆"""
    global STATE

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未收到文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '文件名为空'}), 400

    # 检查文件类型
    fname = file.filename.lower()
    if not (fname.endswith('.xlsx') or fname.endswith('.xls')):
        return jsonify({
            'success': False,
            'error': '不支持的文件格式，请上传 .xlsx 或 .xls 文件'
        }), 400

    # 确保目录存在
    TESTCASES_DIR.mkdir(parents=True, exist_ok=True)

    # 先保存到临时位置校验
    import tempfile
    tmp_path = None
    try:
        suffix = '.xlsx' if fname.endswith('.xlsx') else '.xls'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        # 校验格式
        valid, errmsg = validate_excel(Path(tmp_path))
        if not valid:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
            return jsonify({'success': False, 'error': errmsg}), 400

        # 保存到 testcases 下，用原始文件名
        import shutil
        dest_path = TESTCASES_DIR / file.filename
        shutil.move(tmp_path, str(dest_path))
        tmp_path = None  # 已移动，不需要清理

        # 写入记忆
        write_memory(file.filename)

        # 重新解析
        loaded = read_all_sheets(dest_path)
        _apply_loaded_data(dest_path, loaded)

        return jsonify({
            'success': True,
            'filename': file.filename,
            'total': len(STATE['testcases']),
            'sheet_count': len(STATE.get('sheets', [])),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'处理文件时出错: {str(e)}'}), 500
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@app.route('/api/check-results')
def api_check_results():
    """检查当前活跃文件是否已有测试结果（用于上传前确认）"""
    has = has_test_results()
    display_name = read_memory() or STATE.get('filename', '')
    return jsonify({
        'has_results': has,
        'filename': display_name,
    })


# ============================================================
# 4. HTML 模板（带搜索筛选 + 汇总页的完整前端）
# ============================================================
HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>测试用例记录表</title>
<style>
/* ========== CSS 变量 & 全局 ========== */
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}

:root{
    --bg:#f5f7fa;
    --surface:#ffffff;
    --surface-elevated:#ffffff;
    --text:#111827;
    --text-secondary:#6b7280;
    --text-tertiary:#9ca3af;
    --border:#e5e7eb;
    --border-light:#f3f4f6;
    --primary:#2563eb;
    --primary-light:#eff6ff;
    --primary-dark:#1d4ed8;
    --pass:#16a34a;
    --pass-light:#dcfce7;
    --pass-bg:#f0fdf4;
    --fail:#dc2626;
    --fail-light:#fee2e2;
    --fail-bg:#fef2f2;
    --block:#f59e0b;
    --block-light:#fef3c7;
    --block-bg:#fffbeb;
    --skip:#6b7280;
    --skip-light:#f3f4f6;
    --skip-bg:#f9fafb;
    --purpose-bg:#fffbeb;
    --purpose-border:#f59e0b;
    --shadow:0 1px 2px 0 rgba(0,0,0,0.05);
    --shadow-md:0 4px 6px -1px rgba(0,0,0,0.07),0 2px 4px -2px rgba(0,0,0,0.05);
    --shadow-lg:0 10px 15px -3px rgba(0,0,0,0.08),0 4px 6px -4px rgba(0,0,0,0.03);
    --radius:16px;
    --radius-sm:10px;
    --radius-xs:6px;
}

@media (prefers-color-scheme: dark){
    :root{
        --bg:#0f172a;
        --surface:#1e293b;
        --surface-elevated:#334155;
        --text:#f1f5f9;
        --text-secondary:#94a3b8;
        --text-tertiary:#64748b;
        --border:#334155;
        --border-light:#1e293b;
        --primary:#60a5fa;
        --primary-light:#1e3a8a;
        --primary-dark:#3b82f6;
        --pass:#4ade80;
        --pass-light:#14532d;
        --pass-bg:#14532d;
        --fail:#f87171;
        --fail-light:#7f1d1d;
        --fail-bg:#7f1d1d;
        --block:#fbbf24;
        --block-light:#78350f;
        --block-bg:#78350f;
        --skip:#94a3b8;
        --skip-light:#334155;
        --skip-bg:#334155;
        --purpose-bg:#451a03;
        --purpose-border:#f59e0b;
        --shadow:0 1px 2px 0 rgba(0,0,0,0.3);
        --shadow-md:0 4px 6px -1px rgba(0,0,0,0.4),0 2px 4px -2px rgba(0,0,0,0.3);
        --shadow-lg:0 10px 15px -3px rgba(0,0,0,0.5),0 4px 6px -4px rgba(0,0,0,0.3);
    }
    body{background:var(--bg);}
    .card,.action-card,.summary-stat,.metric-card,.summary-table-wrap,
    .search-result-item,.issue-item,.upload-card,.upload-zone{background:var(--surface);}
    .search-result-item:hover,.issue-item:hover{background:var(--surface-elevated);}
    .field-value,.remark-area,.extra-input,.search-input,.filter-select{color:var(--text);}
    .remark-area,.extra-input,.search-input,.filter-select{background:#0b1120;}
    .btn-nav{background:var(--surface);border-color:var(--border);color:var(--text);}
    .btn-nav:hover{border-color:var(--primary);color:var(--primary);background:var(--primary-light);}
    .summary-table th{background:#0f172a;}
    .summary-table td,.summary-table th{border-color:var(--border);}
    .modal-box{background:var(--surface);}
    .modal-overlay{background:rgba(0,0,0,.6);}
    .state-message .path-hint{background:var(--surface);border-color:var(--border);}
    .kbd{background:#1e293b;border-color:var(--border);}
}

body{
    font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC",
               "Microsoft YaHei","Helvetica Neue",sans-serif;
    background:var(--bg);
    color:var(--text);
    line-height:1.6;
    min-height:100vh;
    -webkit-font-smoothing:antialiased;
    -moz-osx-font-smoothing:grayscale;
}

/* ========== 顶部导航栏 ========== */
.topbar{
    background:var(--surface);
    border-bottom:1px solid var(--border);
    padding:0 20px;
    height:64px;
    display:flex;
    align-items:center;
    justify-content:space-between;
    position:sticky;
    top:0;
    z-index:100;
    box-shadow:var(--shadow);
    gap:16px;
}
.topbar-left{display:flex;align-items:center;gap:12px;white-space:nowrap;}
.topbar-logo{
    width:36px;height:36px;border-radius:var(--radius-xs);
    background:linear-gradient(135deg,var(--primary),var(--primary-dark));
    display:flex;align-items:center;justify-content:center;
    color:#fff;font-size:18px;box-shadow:var(--shadow-md);
}
.topbar-title{font-size:17px;font-weight:700;color:var(--text);letter-spacing:-.3px;}
.topbar-title span{font-size:12px;font-weight:500;color:var(--text-secondary);margin-left:6px;}

/* 视图切换标签 */
.view-tabs{
    display:flex;
    gap:6px;
    background:var(--bg);
    padding:5px;
    border-radius:var(--radius-sm);
    border:1px solid var(--border-light);
}
.view-tab{
    padding:8px 18px;
    border:1px solid transparent;
    background:transparent;
    font-size:13px;
    font-weight:600;
    cursor:pointer;
    transition:all .2s ease;
    font-family:inherit;
    color:var(--text-secondary);
    border-radius:var(--radius-xs);
    display:flex;align-items:center;gap:6px;
}
.view-tab:hover:not(.active){color:var(--text);background:rgba(37,99,235,.06);}
.view-tab.active{
    background:var(--surface);
    color:var(--primary);
    border-color:var(--border);
    box-shadow:var(--shadow);
}

.topbar-right{display:flex;align-items:center;gap:10px;font-size:13px;color:var(--text-secondary);white-space:nowrap;}
#topbarFilename{
    background:var(--bg);
    padding:6px 12px;border-radius:var(--radius-xs);
    border:1px solid var(--border-light);
    font-size:12px;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;
}

/* ========== 搜索栏 ========== */
.search-bar-wrap{
    max-width:960px;
    margin:0 auto;
    padding:16px 24px 0;
    display:flex;
    gap:10px;
    align-items:center;
    flex-wrap:wrap;
}
.search-input{
    flex:1;
    min-width:180px;
    padding:10px 16px;
    border:1px solid var(--border);
    border-radius:var(--radius-sm);
    font-size:14px;
    font-family:inherit;
    background:var(--surface);
    color:var(--text);
    transition:all .2s;
    box-shadow:var(--shadow);
}
.search-input:focus{outline:none;border-color:var(--primary);box-shadow:0 0 0 4px rgba(37,99,235,.12);}
.search-input::placeholder{color:var(--text-tertiary);}

.filter-select{
    padding:10px 14px;
    border:1px solid var(--border);
    border-radius:var(--radius-sm);
    font-size:13px;
    font-family:inherit;
    background:var(--surface);
    color:var(--text);
    cursor:pointer;
    min-width:100px;
    box-shadow:var(--shadow);
    transition:all .2s;
}
.filter-select:focus{outline:none;border-color:var(--primary);box-shadow:0 0 0 4px rgba(37,99,235,.12);}

.search-clear{
    background:var(--surface);
    border:1px solid var(--border);
    border-radius:var(--radius-sm);
    font-size:13px;
    color:var(--text-secondary);
    cursor:pointer;
    white-space:nowrap;
    padding:10px 14px;
    font-family:inherit;
    font-weight:600;
    transition:all .15s;
    box-shadow:var(--shadow);
}
.search-clear:hover{border-color:var(--primary);color:var(--primary);background:var(--primary-light);}

.search-info{
    max-width:960px;
    margin:0 auto;
    padding:6px 24px 0;
    font-size:12px;
    color:var(--text-secondary);
}

/* ========== 进度条 ========== */
.progress-wrap{
    padding:16px 24px 0;
    max-width:960px;
    margin:0 auto;
}
.progress-info{display:flex;justify-content:space-between;margin-bottom:8px;font-size:13px;color:var(--text-secondary);font-weight:500;}
.progress-bar{height:8px;background:var(--border-light);border-radius:4px;overflow:hidden;}
.progress-fill{height:100%;background:linear-gradient(90deg,var(--primary),var(--primary-dark));border-radius:4px;transition:width .5s ease;}
.progress-fill.done{background:linear-gradient(90deg,var(--pass),#22c55e);}
.legend{display:flex;gap:16px;margin-top:8px;font-size:12px;color:var(--text-secondary);flex-wrap:wrap;}
.legend-dot{display:inline-block;width:10px;height:10px;border-radius:50%;margin-right:4px;vertical-align:-1px;}
.legend-dot.pass{background:var(--pass)}
.legend-dot.fail{background:var(--fail)}
.legend-dot.block{background:var(--block)}
.legend-dot.skip{background:var(--skip)}

/* ========== 主内容区 ========== */
.main-container{max-width:960px;margin:0 auto;padding:20px 24px 40px;}

/* 空状态 */
.state-message{text-align:center;padding:80px 20px;color:var(--text-secondary);}
.state-message .icon{
    font-size:52px;margin-bottom:20px;
    width:90px;height:90px;line-height:90px;border-radius:50%;
    background:var(--surface);box-shadow:var(--shadow-md);
    display:inline-flex;align-items:center;justify-content:center;
}
.state-message h2{font-size:22px;margin-bottom:10px;color:var(--text);font-weight:700;}
.state-message p{font-size:14px;max-width:420px;margin:0 auto;line-height:1.7;}
.state-message .path-hint{
    display:inline-block;background:var(--surface);border:1px dashed var(--border);
    border-radius:var(--radius-sm);padding:8px 16px;margin-top:14px;
    font-family:"SF Mono","Fira Code",monospace;font-size:13px;
    box-shadow:var(--shadow);
}

/* ========== 用例卡片 ========== */
.card{background:var(--surface);border-radius:var(--radius);box-shadow:var(--shadow-md);overflow:hidden;margin-bottom:16px;border:1px solid var(--border-light);}
.card-header{
    padding:18px 24px 16px;border-bottom:1px solid var(--border-light);
    display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:10px;
}
.card-navigator{font-size:14px;color:var(--text-secondary);white-space:nowrap;display:flex;align-items:center;gap:8px;}
.card-navigator strong{color:var(--primary);font-size:18px;font-weight:800;}
.card-body{padding:8px 24px 22px;}

/* 结果状态徽标 */
.result-badge{
    display:inline-flex;align-items:center;gap:5px;padding:4px 12px;border-radius:20px;font-size:12px;font-weight:700;
}
.result-badge-pass{background:var(--pass-light);color:var(--pass)}
.result-badge-fail{background:var(--fail-light);color:var(--fail)}
.result-badge-block{background:var(--block-light);color:var(--block)}
.result-badge-skip{background:var(--skip-light);color:var(--skip)}

/* 字段展示 */
.field-row{padding:14px 0;border-bottom:1px solid var(--border-light);transition:all .2s;}
.field-row:last-child{border-bottom:none;}
.field-label{font-size:12px;font-weight:700;color:var(--text-secondary);margin-bottom:6px;text-transform:uppercase;letter-spacing:.4px;}
.field-value{font-size:15px;color:var(--text);line-height:1.75;white-space:pre-wrap;word-break:break-word;}

/* 字段通用块：左侧彩色竖线 + 浅色背景 + 圆角 + 悬停动效 */
.field-section .field-value,
.field-id .field-value,
.field-title .field-value{
    padding:14px 18px;border-radius:var(--radius-sm);border-left:4px solid;
    transition:all .2s ease;
}

/* 通用字段：中性灰色 */
.field-section .field-value{
    color:var(--text);background:var(--bg);border-left-color:#9ca3af;
}

/* 编号 */
.field-id .field-value{
    font-family:"SF Mono","Fira Code","Consolas",monospace;font-size:13px;
    display:inline-flex;align-items:center;padding:8px 16px;border-radius:var(--radius-sm);
    color:var(--primary);background:var(--primary-light);border-left-color:var(--primary);font-weight:700;
    box-shadow:var(--shadow);
}

/* 标题 */
.field-title .field-value{
    font-size:20px;font-weight:800;line-height:1.4;letter-spacing:-.3px;
    background:var(--surface);border-left-color:var(--primary);border:1px solid var(--border-light);
    box-shadow:var(--shadow);
}

/* 前置条件 */
.field-section-pre .field-value{
    color:#1e40af;background:#eff6ff;border-left-color:#3b82f6;
}

/* 测试步骤 */
.field-steps .field-value{
    color:#5b21b6;background:#f5f3ff;border-left-color:#8b5cf6;
}

/* 预期结果 */
.field-expected .field-value{
    color:#15803d;background:var(--pass-bg);border-left-color:#22c55e;
}

/* 测试目的 */
.field-purpose .field-value{
    color:#92400e;background:var(--purpose-bg);border-left-color:var(--purpose-border);
}

/* 字段块悬停效果 */
.field-section:hover .field-value,
.field-title:hover .field-value{
    transform:translateY(-1px);box-shadow:var(--shadow-md);
}
.field-section .field-value:hover,
.field-section:hover .field-value{background:#f3f4f6;}
.field-section-pre .field-value:hover,
.field-section-pre:hover .field-value{background:#dbeafe;}
.field-steps .field-value:hover,
.field-steps:hover .field-value{background:#ede9fe;}
.field-expected .field-value:hover,
.field-expected:hover .field-value{background:#dcfce7;}
.field-purpose .field-value:hover,
.field-purpose:hover .field-value{background:#fef3c7;}
.field-title:hover .field-value{background:var(--surface);}
.field-id:hover .field-value{transform:translateY(-1px);box-shadow:var(--shadow-md);}

/* 元信息标签：优先级、模块、分类 */
.field-meta{display:inline-flex;align-items:center;gap:16px;padding:6px 0;border-bottom:none;}
.field-meta .field-label{margin-bottom:0;}
.field-meta .field-value{
    font-size:13px;background:var(--bg);padding:5px 14px;border-radius:20px;font-weight:600;
    border:1px solid var(--border);border-left:none;box-shadow:var(--shadow);
}

/* ========== 操作区 ========== */
.action-card{background:var(--surface);border-radius:var(--radius);box-shadow:var(--shadow-md);padding:22px 24px 24px;margin-bottom:16px;border:1px solid var(--border-light);}
.action-card h3{font-size:14px;margin-bottom:16px;color:var(--text);font-weight:700;letter-spacing:.3px;}

.result-group{display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap;}
.result-btn{
    flex:1;min-width:90px;padding:14px 8px;border:2px solid var(--border);border-radius:var(--radius-sm);
    background:var(--surface);font-size:14px;font-weight:700;cursor:pointer;
    transition:all .2s ease;text-align:center;color:var(--text);user-select:none;
    display:flex;flex-direction:column;align-items:center;gap:4px;
}
.result-btn:hover{transform:translateY(-2px);box-shadow:var(--shadow-md);border-color:var(--primary);color:var(--primary);}
.result-btn:active{transform:scale(.97);}

.result-btn.selected-pass{border-color:var(--pass);background:var(--pass-bg);color:var(--pass);box-shadow:0 0 0 4px rgba(22,163,74,.12);}
.result-btn.selected-fail{border-color:var(--fail);background:var(--fail-bg);color:var(--fail);box-shadow:0 0 0 4px rgba(220,38,38,.12);}
.result-btn.selected-block{border-color:var(--block);background:var(--block-bg);color:var(--block);box-shadow:0 0 0 4px rgba(245,158,11,.12);}
.result-btn.selected-skip{border-color:var(--skip);background:var(--skip-bg);color:var(--skip);box-shadow:0 0 0 4px rgba(107,114,128,.12);}

.result-emoji{font-size:20px;display:block;margin-bottom:2px;}

.remark-area{
    width:100%;min-height:86px;padding:14px 16px;border:1px solid var(--border);
    border-radius:var(--radius-sm);font-size:14px;font-family:inherit;line-height:1.7;
    resize:vertical;transition:all .2s;color:var(--text);background:var(--bg);
}
.remark-area:focus{outline:none;border-color:var(--primary);background:var(--surface);box-shadow:0 0 0 4px rgba(37,99,235,.12);}
.remark-area::placeholder{color:var(--text-tertiary);}

.btn{
    display:inline-flex;align-items:center;justify-content:center;gap:6px;padding:10px 22px;
    border:none;border-radius:var(--radius-sm);font-size:14px;font-weight:700;
    cursor:pointer;transition:all .2s ease;font-family:inherit;user-select:none;
}
.btn:active{transform:scale(.97);}

.btn-save{background:linear-gradient(135deg,var(--primary),var(--primary-dark));color:#fff;padding:13px 30px;font-size:15px;box-shadow:0 4px 10px rgba(37,99,235,.35);}
.btn-save:hover{background:linear-gradient(135deg,var(--primary-dark),#1e40af);box-shadow:0 6px 14px rgba(37,99,235,.4);transform:translateY(-1px);}
.btn-save:disabled{background:#a5b4fc;cursor:not-allowed;box-shadow:none;transform:none;}

.btn-nav{background:var(--surface);border:1px solid var(--border);color:var(--text);padding:11px 24px;font-size:14px;box-shadow:var(--shadow);}
.btn-nav:hover{border-color:var(--primary);color:var(--primary);background:var(--primary-light);}
.btn-nav:disabled{opacity:.45;cursor:not-allowed;box-shadow:none;}
.btn-nav:disabled:hover{border-color:var(--border);color:var(--text);background:var(--surface);}

.btn-sm{padding:8px 16px;font-size:12px;}
.btn-outline{background:var(--surface);border:1px solid var(--border);color:var(--text);box-shadow:var(--shadow);}
.btn-outline:hover{border-color:var(--primary);color:var(--primary);background:var(--primary-light);}

/* ========== 额外字段 ========== */
.extra-label {
    display: block; font-size: 12px; font-weight: 700;
    color: var(--text-secondary); margin-bottom: 6px;
}
.extra-input {
    width: 100%; padding: 10px 12px; border: 1px solid var(--border);
    border-radius: var(--radius-sm); font-size: 13px; font-family: inherit;
    background: var(--bg); color: var(--text);
    transition: all .2s; box-sizing: border-box;
}
.extra-input:focus { outline: none; border-color: var(--primary); background: var(--surface); box-shadow: 0 0 0 4px rgba(37,99,235,.12); }

/* ========== 前置条件美化 ========== */
.field-section-pre .field-value {
    color: #1e40af;
    background: #eff6ff;
    padding: 14px 18px;
    border-radius: var(--radius-sm);
    border-left: 4px solid #3b82f6;
}

/* ========== 结果按钮二次确认 ========== */
.result-btn.selected-pass { background: var(--pass-bg); border-color: #22c55e; color: #15803d; }
.result-btn.selected-fail { background: var(--fail-bg); border-color: #ef4444; color: #b91c1c; }
.result-btn.selected-block { background: var(--block-bg); border-color: #f59e0b; color: #b45309; }
.result-btn.selected-skip { background: var(--skip-bg); border-color: #9ca3af; color: #4b5563; }

/* ========== 问题列表 ========== */
.issue-list { display: flex; flex-direction: column; gap: 12px; }
.issue-item {
    background: var(--surface); border: 1px solid var(--border-light); border-radius: var(--radius);
    padding: 16px 20px; cursor: pointer; transition: all .2s; box-shadow: var(--shadow);
}
.issue-item:hover { border-color: var(--primary); box-shadow: var(--shadow-md); transform: translateY(-2px); }
.issue-item .issue-header { display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 10px; }
.issue-item .issue-id { font-family: monospace; font-size: 13px; color: var(--primary); font-weight: 700; }
.issue-item .issue-title { font-size: 16px; font-weight: 700; margin: 6px 0; color: var(--text); }
.issue-item .issue-meta { font-size: 12px; color: var(--text-secondary); display: flex; gap: 14px; flex-wrap: wrap; }

/* ========== 庆祝动画 ========== */
.celebrate-wrap {
    text-align: center; padding: 50px 24px; position: relative;
    min-height: 400px; display: flex; flex-direction: column;
    align-items: center; justify-content: center;
    background: var(--surface); border-radius: var(--radius); box-shadow: var(--shadow-md);
    border: 1px solid var(--border-light);
}
.celebrate-text { font-size: 30px; font-weight: 800; color: var(--text); margin-top: 24px; z-index: 1; }
.celebrate-sub { font-size: 16px; color: var(--text-secondary); margin-top: 10px; z-index: 1; }
.confetti {
    position: absolute; width: 10px; height: 10px; border-radius: 50%;
    animation: confettiFall 3s ease-out forwards;
}
@keyframes confettiFall {
    0% { transform: translate(0,0) rotate(0deg) scale(1); opacity: 1; }
    50% { opacity: 1; }
    100% { transform: translate(var(--dx), var(--dy)) rotate(var(--rot)) scale(0); opacity: 0; }
}

/* ========== 搜索结果列表 ========== */
.search-result-list{display:flex;flex-direction:column;gap:10px;}
.search-result-item{
    background:var(--surface);border:1px solid var(--border-light);border-radius:var(--radius-sm);
    padding:14px 18px;cursor:pointer;transition:all .2s;
    display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap;
    box-shadow:var(--shadow);
}
.search-result-item:hover{border-color:var(--primary);box-shadow:var(--shadow-md);transform:translateY(-2px);}
.search-result-item .sr-left{flex:1;min-width:0;}
.search-result-item .sr-id{font-family:"SF Mono","Fira Code",monospace;font-size:12px;color:var(--primary);font-weight:700;}
.search-result-item .sr-title{font-size:15px;font-weight:700;margin:4px 0;word-break:break-word;color:var(--text);}
.search-result-item .sr-meta{font-size:12px;color:var(--text-secondary);}
.search-result-item .sr-snippet{font-size:12px;color:var(--text-secondary);margin-top:5px;
    white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.search-result-item .sr-snippet em{background:#fde68a;font-style:normal;padding:0 3px;border-radius:3px;}
.search-result-item .sr-right{flex-shrink:0;display:flex;align-items:center;gap:10px;}

/* ========== 底部分页 ========== */
.pagination{display:flex;align-items:center;justify-content:center;gap:10px;margin-top:18px;font-size:14px;}
.pagination .page-info{color:var(--text-secondary);font-weight:500;}

/* ========== 汇总页 ========== */
.summary-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:20px;}
.summary-stat{
    background:var(--surface);border-radius:var(--radius);box-shadow:var(--shadow);
    padding:22px 16px;text-align:center;border:1px solid var(--border-light);
    transition:transform .2s;
}
.summary-stat:hover{transform:translateY(-2px);box-shadow:var(--shadow-md);}
.summary-stat .stat-number{font-size:34px;font-weight:800;line-height:1.1;}
.summary-stat .stat-label{font-size:13px;color:var(--text-secondary);margin-top:6px;font-weight:600;}
.stat-pass .stat-number{color:var(--pass)}
.stat-fail .stat-number{color:var(--fail)}
.stat-block .stat-number{color:var(--block)}
.stat-skip .stat-number{color:var(--skip)}
.stat-total .stat-number{color:var(--primary)}

/* 汇总表格 */
.summary-table-wrap{background:var(--surface);border-radius:var(--radius);box-shadow:var(--shadow);overflow:hidden;margin-bottom:16px;border:1px solid var(--border-light);}
.summary-table-wrap h3{padding:18px 22px 14px;font-size:15px;border-bottom:1px solid var(--border-light);font-weight:700;}
.summary-table{width:100%;border-collapse:collapse;font-size:13px;}
.summary-table th,.summary-table td{padding:12px 16px;text-align:center;border-bottom:1px solid var(--border-light);}
.summary-table th{background:#f9fafb;font-weight:700;font-size:12px;color:var(--text-secondary);text-transform:uppercase;letter-spacing:.3px;}
.summary-table td:first-child{text-align:left;font-weight:700;}
.summary-table .cell-pass{color:var(--pass);font-weight:700;}
.summary-table .cell-fail{color:var(--fail);font-weight:700;}
.summary-table .cell-block{color:var(--block);font-weight:700;}
.summary-table .cell-skip{color:var(--skip);}
.summary-table .bar-cell{width:120px;}
.summary-table .mini-bar{height:7px;background:var(--border-light);border-radius:4px;overflow:hidden;display:inline-block;width:100%;}
.summary-table .mini-bar-fill{height:100%;border-radius:4px;transition:width .5s;}
.mini-bar-fill.pass{background:var(--pass)}
.mini-bar-fill.fail{background:var(--fail)}
.mini-bar-fill.block{background:var(--block)}

/* 汇总指标行 */
.summary-metrics{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:20px;}
.metric-card{
    background:var(--surface);border-radius:var(--radius);box-shadow:var(--shadow);
    padding:18px 22px;display:flex;align-items:center;gap:16px;flex:1;min-width:220px;
    border:1px solid var(--border-light);transition:transform .2s;
}
.metric-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-md);}
.metric-ring{width:72px;height:72px;position:relative;flex-shrink:0;}
.metric-ring svg{transform:rotate(-90deg);}
.metric-ring .bg{fill:none;stroke:var(--border-light);stroke-width:7;}
.metric-ring .fg{fill:none;stroke-width:7;stroke-linecap:round;transition:stroke-dashoffset .6s ease;}
.metric-ring .pct{position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);font-size:17px;font-weight:800;}
.metric-text h4{font-size:14px;margin-bottom:3px;color:var(--text);font-weight:700;}
.metric-text p{font-size:12px;color:var(--text-secondary);}

/* ========== 底部导航 ========== */
.nav-bar{display:flex;align-items:center;justify-content:center;gap:16px;padding:10px 0;}
.case-jump-select{appearance:none;-webkit-appearance:none;padding:8px 32px 8px 14px;font-size:14px;font-weight:600;color:var(--text);background:var(--bg);border:1px solid var(--border);border-radius:var(--radius-sm);cursor:pointer;outline:none;min-width:180px;max-width:260px;text-overflow:ellipsis;white-space:nowrap;overflow:hidden;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24' fill='none' stroke='%2364748b' stroke-width='2'%3E%3Cpath d='M6 9l6 6 6-6'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 10px center;}
.case-jump-select:focus{border-color:var(--primary);box-shadow:0 0 0 2px var(--primary-light);}
.case-jump-select option{padding:6px 10px;font-weight:500;}
.nav-hint{font-size:12px;color:var(--text-tertiary);text-align:center;margin-top:8px;}

/* ========== Toast ========== */
.toast{
    position:fixed;top:20px;left:50%;transform:translateX(-50%);padding:13px 26px;
    border-radius:var(--radius-sm);font-size:14px;font-weight:700;
    z-index:999;animation:toastIn .3s ease,toastOut .3s ease 2s forwards;box-shadow:var(--shadow-lg);
    border:1px solid rgba(255,255,255,.2);
}
.toast-success{background:linear-gradient(135deg,#16a34a,#15803d);color:#fff;}
.toast-error{background:linear-gradient(135deg,#dc2626,#b91c1c);color:#fff;}
.toast-warn{background:linear-gradient(135deg,#f59e0b,#d97706);color:#fff;}

@keyframes toastIn{from{opacity:0;transform:translateX(-50%) translateY(-20px)}to{opacity:1;transform:translateX(-50%) translateY(0)}}
@keyframes toastOut{from{opacity:1}to{opacity:0}}

/* ========== 弹窗 ========== */
.modal-overlay{
    position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.5);
    display:flex;align-items:center;justify-content:center;z-index:200;
    backdrop-filter:blur(4px);
}
.modal-box{
    background:var(--surface);border-radius:var(--radius);padding:28px;max-width:420px;
    width:90%;box-shadow:0 25px 70px rgba(0,0,0,.18);text-align:center;border:1px solid var(--border-light);
}
.modal-box h3{margin-bottom:12px;font-size:19px;font-weight:800;color:var(--text);}
.modal-box p{margin-bottom:24px;color:var(--text-secondary);font-size:14px;line-height:1.7;}
.modal-actions{display:flex;gap:12px;justify-content:center;}
.btn-modal-primary{background:var(--primary);color:#fff;padding:10px 22px;border:none;border-radius:var(--radius-sm);font-size:14px;font-weight:700;cursor:pointer;font-family:inherit;transition:all .2s;}
.btn-modal-primary:hover{background:var(--primary-dark);transform:translateY(-1px);}
.btn-modal-ghost{background:var(--surface);color:var(--text-secondary);padding:10px 22px;border:1px solid var(--border);border-radius:var(--radius-sm);font-size:14px;cursor:pointer;font-family:inherit;font-weight:700;transition:all .2s;}
.btn-modal-ghost:hover{border-color:var(--primary);color:var(--primary);background:var(--primary-light);}

/* ========== 快捷键提示 ========== */
.kbd{display:inline-block;background:var(--bg);border:1px solid var(--border);border-radius:5px;padding:2px 7px;font-family:"SF Mono","Fira Code",monospace;font-size:11px;vertical-align:1px;color:var(--text-secondary);font-weight:700;}

.highlight{background:#fde68a;padding:0 2px;border-radius:3px;}

/* ========== 响应式 ========== */
@media(max-width:760px){
    .topbar{padding:0 14px;height:58px;}
    .topbar-title span{display:none;}
    .view-tab{padding:7px 12px;font-size:12px;}
    .topbar-right #btnReplace span,.topbar-right #btnReload span{display:none;}
    .card-header{padding:14px 16px 12px}
    .card-body{padding:6px 16px 18px}
    .action-card{padding:18px 16px}
    .result-group{gap:8px}
    .result-btn{min-width:70px;padding:12px 4px;font-size:13px}
    .result-emoji{font-size:18px}
    .field-title .field-value{font-size:19px}
    .main-container{padding:14px 14px 40px}
    .progress-wrap{padding:14px 14px 0}
    .search-bar-wrap{padding:14px 14px 0}
    .search-info{padding:4px 14px 0}
    .summary-grid{grid-template-columns:repeat(2,1fr)}
    .summary-metrics{flex-direction:column;}
    .extra-fields{grid-template-columns:1fr!important;}
    .nav-bar{gap:10px;}
}

@media print{
    body{background:#fff}
    .topbar,.action-card,.nav-bar,.nav-hint,.progress-wrap,.search-bar-wrap,.search-info{display:none}
    .card{box-shadow:none;border:1px solid #e5e7eb}
}
</style>
</head>
<body>

<!-- 顶部导航 -->
<div class="topbar" id="topbarMain">
    <div class="topbar-left">
        <span class="topbar-logo">🧪</span>
        <span class="topbar-title">测试用例记录表<span>TestCase Viewer</span></span>
    </div>
    <div class="view-tabs">
        <button class="view-tab active" id="tabExecute" onclick="switchView('execute')">▶ 执行测试</button>
        <button class="view-tab" id="tabSummary" onclick="switchView('summary')">📊 汇总统计</button>
        <button class="view-tab" id="tabIssues" onclick="switchView('issues')">📝 问题列表</button>
    </div>
    <div class="topbar-right">
        <button class="btn btn-outline btn-sm" id="btnReplace" onclick="triggerReplaceFile()" title="上传新的测试用例 Excel 文件"><span>📂</span> 更换文件</button>
        <button class="btn btn-outline btn-sm" id="btnReload" onclick="reloadData()" title="重新加载当前 Excel（Excel 内容更新后点此刷新）"><span>↻</span> 刷新</button>
        <span id="topbarFilename">未加载</span>
    </div>
</div>

<!-- ===== 执行测试视图 ===== -->
<div id="viewExecute">
    <!-- 进度条 -->
    <div class="progress-wrap" id="progressWrap" style="display:none">
        <div class="progress-info">
            <span id="progressText">共 0 条</span>
            <span id="progressPercent">0%</span>
        </div>
        <div class="progress-bar"><div class="progress-fill" id="progressFill" style="width:0%"></div></div>
        <div class="legend">
            <span><span class="legend-dot pass"></span>通过 <span id="countPass">0</span></span>
            <span><span class="legend-dot fail"></span>失败 <span id="countFail">0</span></span>
            <span><span class="legend-dot block"></span>阻塞 <span id="countBlock">0</span></span>
            <span><span class="legend-dot skip"></span>跳过 <span id="countSkip">0</span></span>
            <span>未执行 <span id="countNone">0</span></span>
        </div>
    </div>

    <div class="main-container" id="execContainer"></div>

    <div class="nav-bar" id="navBar" style="display:none">
        <button class="btn btn-nav" id="btnPrev" onclick="goTo(-1)" title="&#8592; 键">&#9664; 上一条</button>
        <select class="case-jump-select" id="caseJumpSelect" onchange="if(this.value!==''){jumpToExecute(parseInt(this.value));}">
            <option value="">&#128269; 选择用例跳转...</option>
        </select>
        <button class="btn btn-nav" id="btnNext" onclick="goTo(1)" title="&#8594; 键">下一条 &#9654;</button>
    </div>
    <div class="nav-hint" id="navHint" style="display:none">
        <span class="kbd">&#8592;</span> 上一条 &nbsp; <span class="kbd">&#8594;</span> 下一条 &nbsp; <span class="kbd">Ctrl+S</span> 保存
    </div>
</div>

<!-- ===== 搜索筛选视图 ===== -->
<div id="viewSearch" style="display:none;">
    <div class="search-bar-wrap" style="padding-top:16px;">
        <input type="text" class="search-input" id="searchKeyword"
               placeholder="输入关键词搜索（用例编号、标题、步骤等）..."
               oninput="doSearch()" onkeydown="if(event.key==='Enter')doSearch()">
        <select class="filter-select" id="filterResult" onchange="doSearch()">
            <option value="">全部结果</option>
            <option value="未执行">未执行</option>
            <option value="通过">通过</option>
            <option value="失败">失败</option>
            <option value="阻塞">阻塞</option>
            <option value="跳过">跳过</option>
        </select>
        <select class="filter-select" id="filterPriority" onchange="doSearch()">
            <option value="">全部优先级</option>
        </select>
        <select class="filter-select" id="filterModule" onchange="doSearch()">
            <option value="">全部模块</option>
        </select>
        <button class="search-clear" onclick="clearSearch()">&#10005; 清除</button>
    </div>
    <div class="search-info" id="searchInfo"></div>

    <div class="main-container" id="searchContainer">
        <div class="state-message"><div class="icon">&#128269;</div><h2>输入关键词开始搜索</h2></div>
    </div>

    <div class="pagination" id="searchPagination" style="display:none;"></div>
</div>

<!-- ===== 汇总统计视图 ===== -->
<div id="viewSummary" style="display:none;">
    <div class="main-container" id="summaryContainer">
        <div class="state-message"><div class="icon">&#9203;</div><h2>加载中...</h2></div>
    </div>
</div>

<!-- ===== 问题列表视图 ===== -->
<div id="viewIssues" style="display:none;">
    <div class="main-container" id="issuesContainer">
        <div class="state-message"><div class="icon">&#9203;</div><h2>加载中...</h2></div>
    </div>
</div>

<!-- Toast -->
<div id="toastContainer"></div>

<script>
// ============================================================
// 全局状态
// ============================================================
let currentView = 'execute';    // execute | search | summary
let currentIndex = 0;
let totalCount = 0;
let dirty = false;
let selectedResult = '';
let allStatuses = [];

let searchPage = 0;
let searchPageSize = 30;
let searchLastParams = null;

// ============================================================
// 视图切换
// ============================================================
function switchView(view) {
    if (dirty && view !== 'execute') {
        showUnsavedDialog().then(confirmed => {
            if (confirmed) _doSwitch(view);
        });
        return;
    }
    _doSwitch(view);
}

function _doSwitch(view) {
    currentView = view;

    document.getElementById('viewExecute').style.display = view === 'execute' ? 'block' : 'none';
    document.getElementById('viewSearch').style.display = view === 'search' ? 'block' : 'none';
    document.getElementById('viewSummary').style.display = view === 'summary' ? 'block' : 'none';
    var issuesDiv = document.getElementById('viewIssues');
    if (issuesDiv) issuesDiv.style.display = view === 'issues' ? 'block' : 'none';

    document.getElementById('tabExecute').classList.toggle('active', view === 'execute');
    document.getElementById('tabSearch').classList.toggle('active', view === 'search');
    document.getElementById('tabSummary').classList.toggle('active', view === 'summary');
    var tabIssues = document.getElementById('tabIssues');
    if (tabIssues) tabIssues.classList.toggle('active', view === 'issues');

    if (view === 'search') {
        loadFilterOptions();
        searchPage = 0;
        doSearch();
    } else if (view === 'summary') {
        loadSummary();
    } else if (view === 'issues') {
        loadIssues();
    }

    window.scrollTo({top: 0, behavior: 'smooth'});
}

// ============================================================
// 初始化
// ============================================================
async function init() {
    try {
        const resp = await fetch('/api/init');
        const data = await resp.json();

        if (!data.loaded) {
            // 没有活跃文件 → 显示上传页
            if (data.has_active) {
                // 有文件但读不到数据（可能是空 Excel）
                showEmptyState();
            } else {
                showUploadPage();
            }
            return;
        }

        totalCount = data.total;
        document.getElementById('topbarFilename').textContent =
            '📂 ' + (data.filename || '已加载');
        document.getElementById('progressWrap').style.display = 'block';
        document.getElementById('navBar').style.display = 'flex';
        document.getElementById('navHint').style.display = 'block';

        await refreshStatuses();
        await loadCaseTitles();
        await loadTestCase(0);
    } catch (err) {
        showEmptyState();
        console.error('初始化失败:', err);
    }
}

function showEmptyState() {
    const container = document.getElementById('execContainer');
    container.innerHTML = `
        <div class="state-message">
            <div class="icon">&#128203;</div>
            <h2>未读取到用例数据</h2>
            <p>当前活跃文件中未检测到有效用例数据。<br>请确保第一行为表头，从第二行开始为用例数据。<br>你也可以重新上传一个新的测试用例文件。</p>
            <button class="btn btn-save" style="margin-top:16px;" onclick="showReplaceDialog()">
                &#128194; 上传新文件
            </button>
        </div>
    `;
}

// ============================================================
// 上传页
// ============================================================
function showUploadPage() {
    const container = document.getElementById('execContainer');
    document.getElementById('progressWrap').style.display = 'none';
    document.getElementById('navBar').style.display = 'none';
    document.getElementById('navHint').style.display = 'none';

    container.innerHTML = `
        <div class="upload-page" id="uploadPage">
            <div class="upload-card" id="uploadCard">
                <div class="upload-icon">🚀</div>
                <h2>欢迎使用测试用例记录表</h2>
                <p class="upload-desc">请上传你的测试用例 Excel 文件开始使用<br>支持 .xlsx 和 .xls 格式</p>
                <div class="upload-zone" id="uploadZone">
                    <div class="upload-zone-icon">📂</div>
                    <div class="upload-zone-text">拖拽 Excel 文件到此处</div>
                    <div class="upload-zone-hint">或点击下方按钮选择文件</div>
                    <input type="file" id="fileInput" accept=".xlsx,.xls" style="display:none"
                           onchange="handleFileSelect(this)">
                    <button class="btn btn-save" onclick="document.getElementById('fileInput').click()"
                            style="margin-top:16px;">
                        📁 选择文件
                    </button>
                </div>
                <div class="upload-status" id="uploadStatus" style="display:none;"></div>
            </div>
        </div>
        <style>
            .upload-page {
                display: flex; align-items: center; justify-content: center;
                min-height: 70vh; padding: 40px 20px;
            }
            .upload-card {
                background: var(--surface); border-radius: var(--radius);
                box-shadow: var(--shadow-lg); padding: 44px 40px;
                max-width: 520px; width: 100%; text-align: center;
                border: 1px solid var(--border-light);
            }
            .upload-icon { font-size: 56px; margin-bottom: 16px; }
            .upload-card h2 { font-size: 24px; margin-bottom: 10px; color: var(--text); font-weight: 800; }
            .upload-desc { font-size: 14px; color: var(--text-secondary); margin-bottom: 32px; line-height: 1.7; }
            .upload-zone {
                border: 2px dashed var(--border); border-radius: var(--radius);
                padding: 36px 24px; transition: all .25s;
                cursor: pointer; background: var(--bg);
            }
            .upload-zone:hover, .upload-zone.drag-over {
                border-color: var(--primary); background: var(--primary-light);
                transform: translateY(-2px); box-shadow: var(--shadow-md);
            }
            .upload-zone-icon { font-size: 44px; margin-bottom: 10px; }
            .upload-zone-text { font-size: 17px; font-weight: 700; color: var(--text); margin-bottom: 6px; }
            .upload-zone-hint { font-size: 13px; color: var(--text-secondary); }
            .upload-status { margin-top: 18px; padding: 12px; border-radius: var(--radius-sm); font-size: 14px; font-weight: 600; }
            .upload-status.loading { background: #eff6ff; color: #1d4ed8; }
            .upload-status.error { background: #fef2f2; color: #991b1b; }
        </style>
    `;

    // 拖拽事件
    const zone = document.getElementById('uploadZone');
    if (zone) {
        zone.addEventListener('dragover', (e) => { e.preventDefault(); zone.classList.add('drag-over'); });
        zone.addEventListener('dragleave', () => { zone.classList.remove('drag-over'); });
        zone.addEventListener('drop', (e) => {
            e.preventDefault();
            zone.classList.remove('drag-over');
            const files = e.dataTransfer.files;
            if (files.length > 0) {
                uploadFile(files[0]);
            }
        });
    }
}

async function handleFileSelect(input) {
    if (input.files.length > 0) {
        await uploadFile(input.files[0]);
    }
}

async function uploadFile(file) {
    // 检查文件类型
    const name = file.name.toLowerCase();
    if (!name.endsWith('.xlsx') && !name.endsWith('.xls')) {
        showUploadStatus('不支持的文件格式，请上传 .xlsx 或 .xls 文件', 'error');
        return;
    }

    // 检查是否已有测试结果 → 二次确认
    try {
        const checkResp = await fetch('/api/check-results');
        const checkData = await checkResp.json();
        if (checkData.has_results) {
            const confirmed = await showReplaceConfirmDialog(checkData.filename);
            if (!confirmed) return;
        }
    } catch (err) {
        // 检查失败不影响上传
    }

    showUploadStatus('⏳ 正在上传并解析文件...', 'loading');

    const formData = new FormData();
    formData.append('file', file);

    try {
        const resp = await fetch('/api/upload', { method: 'POST', body: formData });
        const data = await resp.json();

        if (data.success) {
            showUploadStatus('✅ 上传成功！共 ' + data.total + ' 条用例', '');
            totalCount = data.total;
            document.getElementById('topbarFilename').textContent =
                '📂 ' + (data.filename || '已加载');
            document.getElementById('progressWrap').style.display = 'block';
            document.getElementById('navBar').style.display = 'flex';
            document.getElementById('navHint').style.display = 'block';

            // 短暂延迟后切换到执行页
            setTimeout(async () => {
                await refreshStatuses();
                await loadCaseTitles();
                await loadTestCase(0);
            }, 600);
        } else {
            showUploadStatus('❌ ' + (data.error || '上传失败'), 'error');
        }
    } catch (err) {
        showUploadStatus('❌ 网络错误: ' + err.message, 'error');
    }
}

function showUploadStatus(msg, type) {
    const el = document.getElementById('uploadStatus');
    if (!el) return;
    el.style.display = 'block';
    el.textContent = msg;
    el.className = 'upload-status ' + type;
}

// ============================================================
// 更换文件对话框
// ============================================================
function showReplaceDialog() {
    // 构造一个隐藏的 file input
    const input = document.createElement('input');
    input.type = 'file';
    input.accept = '.xlsx,.xls';
    input.onchange = async () => {
        if (input.files.length > 0) {
            // 检查是否有结果
            try {
                const checkResp = await fetch('/api/check-results');
                const checkData = await checkResp.json();
                if (checkData.has_results) {
                    const confirmed = await showReplaceConfirmDialog(checkData.filename);
                    if (!confirmed) return;
                }
            } catch (err) {}
            await uploadFile(input.files[0]);
        }
    };
    input.click();
}

function showReplaceConfirmDialog(filename) {
    return new Promise((resolve) => {
        const overlay = document.createElement('div');
        overlay.className = 'modal-overlay';
        overlay.innerHTML = `
            <div class="modal-box">
                <h3>⚠️ 确认替换文件</h3>
                <p>当前活跃文件 <strong>${escapeHtml(filename)}</strong> 中已有测试执行结果。<br><br>
                   上传新文件后将覆盖当前文件，<br><strong>已有执行结果将会丢失</strong>。<br>
                   如需保留，请先备份原文件。</p>
                <div class="modal-actions">
                    <button class="btn-modal-ghost" id="modalCancel">取消</button>
                    <button class="btn-modal-primary" id="modalConfirm"
                            style="background:#dc2626;">确认替换</button>
                </div>
            </div>`;
        document.body.appendChild(overlay);

        overlay.querySelector('#modalCancel').onclick = () => {
            document.body.removeChild(overlay);
            resolve(false);
        };
        overlay.querySelector('#modalConfirm').onclick = () => {
            document.body.removeChild(overlay);
            resolve(true);
        };
        overlay.addEventListener('click', (e) => {
            if (e.target === overlay) { document.body.removeChild(overlay); resolve(false); }
        });
    });
}

// 全局都可以触发更换文件（从 topbar 按钮）
async function triggerReplaceFile() {
    if (dirty) {
        const confirmed = await showUnsavedDialog();
        if (!confirmed) return;
    }
    showReplaceDialog();
}

// ============================================================
// 执行测试 - 加载用例
// ============================================================
async function loadTestCase(index) {
    if (dirty) {
        const confirmed = await showUnsavedDialog();
        if (!confirmed) return;
    }

    try {
        const resp = await fetch('/api/testcase/' + index);
        if (!resp.ok) throw new Error('加载失败');

        const tc = await resp.json();
        currentIndex = index;
        selectedResult = tc._saved_result || '';
        renderTestCase(tc);
        updateNavButtons();
        syncCaseJumpSelect(index);
        dirty = false;
        window.scrollTo({top: 0, behavior: 'smooth'});
    } catch (err) {
        showToast('加载用例失败: ' + err.message, 'error');
    }
}

function renderTestCase(tc) {
    const container = document.getElementById('execContainer');
    const fields = tc._display_fields || [];
    const total = tc._total || totalCount;

    let fieldsHtml = '';
    for (const f of fields) {
        let valueHtml = escapeHtml(f.value);
        if (f.css_class.includes('field-steps')) {
            valueHtml = formatSteps(f.value);
        }
        fieldsHtml += `
            <div class="field-row ${f.css_class}">
                <div class="field-label">${f.label}</div>
                <div class="field-value">${valueHtml}</div>
            </div>`;
    }

    const savedBadge = tc._saved_result
        ? `<span class="result-badge result-badge-${resultCss(tc._saved_result)}">已执行: ${escapeHtml(tc._saved_result)}</span>`
        : '';

    container.innerHTML = `
        <div class="card">
            <div class="card-header">
                <div class="card-navigator">
                    第 <strong>${tc._index + 1}</strong> 条 / 共 ${total} 条
                </div>
                ${savedBadge}
            </div>
            <div class="card-body">${fieldsHtml}</div>
        </div>

        <div class="action-card">
            <h3>&#9989; 测试结果</h3>
            <div class="result-group">
                <button class="result-btn ${selectedResult === '通过' ? 'selected-pass' : ''}"
                        onclick="selectResult('通过', this)">
                    <span class="result-emoji">&#9989;</span>通过
                </button>
                <button class="result-btn ${selectedResult === '失败' ? 'selected-fail' : ''}"
                        onclick="selectResult('失败', this)">
                    <span class="result-emoji">&#10060;</span>失败
                </button>
                <button class="result-btn ${selectedResult === '阻塞' ? 'selected-block' : ''}"
                        onclick="selectResult('阻塞', this)">
                    <span class="result-emoji">&#128683;</span>阻塞
                </button>
                <button class="result-btn ${selectedResult === '跳过' ? 'selected-skip' : ''}"
                        onclick="selectResult('跳过', this)">
                    <span class="result-emoji">&#9193;</span>跳过
                </button>
            </div>

            <h3 style="margin-bottom:8px">&#128221; 实际结果</h3>
            <textarea class="remark-area" id="remarkInput"
                      placeholder="请描述测试执行过程中的实际现象、发现的问题等..."
                      oninput="markDirty()">${escapeHtml(tc._saved_actual_result || '')}</textarea>

            <!-- 新增字段 -->
            <div class="extra-fields" style="margin-top:14px;display:grid;grid-template-columns:1fr 1fr;gap:10px;">
                <div>
                    <label class="extra-label">&#128100; 测试人员</label>
                    <input type="text" class="extra-input" id="inputTester" placeholder="填写测试人员姓名"
                           value="${escapeHtml(tc._saved_tester || '')}" oninput="markDirty()">
                </div>
                <div>
                    <label class="extra-label">&#128030; BugID</label>
                    <input type="text" class="extra-input" id="inputBugId" placeholder="如 PROJ-001"
                           value="${escapeHtml(tc._saved_bug_id || '')}" oninput="markDirty()">
                </div>
                <div>
                    <label class="extra-label">&#128257; Bug频率</label>
                    <select class="extra-input" id="inputBugFreq" onchange="markDirty()">
                        <option value="">-- 请选择 --</option>
                        <option value="必现" ${tc._saved_bug_frequency === '必现' ? 'selected' : ''}>必现</option>
                        <option value="高频" ${tc._saved_bug_frequency === '高频' ? 'selected' : ''}>高频</option>
                        <option value="偶现" ${tc._saved_bug_frequency === '偶现' ? 'selected' : ''}>偶现</option>
                        <option value="1次" ${tc._saved_bug_frequency === '1次' ? 'selected' : ''}>1次</option>
                    </select>
                </div>
                <div>
                    <label class="extra-label">&#128339; 问题时间</label>
                    <input type="text" class="extra-input" id="inputIssueTime" placeholder="选择失败时自动记录"
                           value="${escapeHtml(tc._saved_issue_time || '')}" oninput="markDirty()">
                </div>
            </div>

            <div style="margin-top:14px;display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
                <button class="btn btn-save" id="btnSave" onclick="saveResult()">
                    &#128190; 确认并保存
                </button>
                <span id="saveStatus" style="font-size:13px;color:var(--text-secondary);"></span>
            </div>
        </div>`;

    document.getElementById('remarkInput').addEventListener('keydown', function(e) {
        if ((e.ctrlKey || e.metaKey) && e.key === 's') {
            e.preventDefault();
            saveResult();
        }
    });
}

function formatSteps(text) {
    const lines = text.split('\n').filter(l => l.trim() !== '');
    if (lines.length <= 1) return escapeHtml(text).replace(/\n/g, '<br>');
    let html = '<ol style="padding-left:20px;margin:0;">';
    for (const line of lines) {
        html += `<li style="margin-bottom:4px;">${escapeHtml(line.trim())}</li>`;
    }
    html += '</ol>';
    return html;
}

function selectResult(value, btnEl) {
    selectedResult = value;
    markDirty();
    document.querySelectorAll('.result-btn').forEach(b => b.className = 'result-btn');
    const classMap = {'通过':'selected-pass','失败':'selected-fail','阻塞':'selected-block','跳过':'selected-skip'};
    btnEl.className = 'result-btn ' + (classMap[value] || '');

    // 失败时自动记录问题时间
    if (value === '失败') {
        const now = new Date();
        const timeStr = now.getFullYear() + '-' +
            String(now.getMonth()+1).padStart(2,'0') + '-' +
            String(now.getDate()).padStart(2,'0') + ' ' +
            String(now.getHours()).padStart(2,'0') + ':' +
            String(now.getMinutes()).padStart(2,'0') + ':' +
            String(now.getSeconds()).padStart(2,'0');
        const issueInput = document.getElementById('inputIssueTime');
        if (issueInput) issueInput.value = timeStr;
    } else {
        // 非失败时清空问题时间
        const issueInput = document.getElementById('inputIssueTime');
        if (issueInput) issueInput.value = '';
    }
}

function resultCss(r) {
    const m = {'通过':'pass','失败':'fail','阻塞':'block','跳过':'skip'};
    return m[r] || '';
}

function markDirty() { dirty = true; }

// ============================================================
// 保存结果
// ============================================================
async function saveResult() {
    if (!selectedResult) {
        showToast('请先选择测试结果', 'warn');
        return;
    }
    const actualResult = document.getElementById('remarkInput').value.trim();
    const tester = document.getElementById('inputTester') ? document.getElementById('inputTester').value.trim() : '';
    const bugId = document.getElementById('inputBugId') ? document.getElementById('inputBugId').value.trim() : '';
    const bugFreq = document.getElementById('inputBugFreq') ? document.getElementById('inputBugFreq').value : '';
    const issueTime = document.getElementById('inputIssueTime') ? document.getElementById('inputIssueTime').value.trim() : '';
    const btnSave = document.getElementById('btnSave');
    const saveStatus = document.getElementById('saveStatus');
    btnSave.disabled = true;
    btnSave.textContent = '⏳ 保存中...';

    try {
        const resp = await fetch('/api/save', {
            method:'POST',
            headers:{'Content-Type':'application/json'},
            body:JSON.stringify({
                index:currentIndex, result:selectedResult,
                actual_result: actualResult,
                tester: tester, bug_id: bugId,
                bug_frequency: bugFreq, issue_time: issueTime,
            }),
        });
        const data = await resp.json();
        if (data.success) {
            dirty = false;
            showToast('✅ 已保存！测试结果已写入 Excel', 'success');
            saveStatus.textContent = '✅ 已保存 ' + new Date().toLocaleTimeString('zh-CN');
            await refreshStatuses();
            allStatuses[currentIndex] = selectedResult;
            // 保存后自动跳转到下一条
            const nextIndex = currentIndex + 1;
            if (nextIndex < totalCount) {
                await loadTestCase(nextIndex);
            } else {
                showToast('🎉 已到达最后一条用例', 'success');
            }
        } else {
            showToast('❌ 保存失败: ' + data.error, 'error');
        }
    } catch (err) {
        showToast('❌ 网络错误: ' + err.message, 'error');
    } finally {
        btnSave.disabled = false;
        btnSave.textContent = '💾 确认并保存';
    }
}

// ============================================================
// 导航
// ============================================================
async function goTo(delta) {
    if (dirty) {
        const confirmed = await showUnsavedDialog();
        if (!confirmed) return;
    }
    const newIndex = currentIndex + delta;
    if (newIndex < 0 || newIndex >= totalCount) return;
    loadTestCase(newIndex);
}

function updateNavButtons() {
    document.getElementById('btnPrev').disabled = currentIndex <= 0;
    document.getElementById('btnNext').disabled = currentIndex >= totalCount - 1;
}

async function loadCaseTitles() {
    try {
        const resp = await fetch('/api/titles');
        if (!resp.ok) return;
        const titles = await resp.json();
        const sel = document.getElementById('caseJumpSelect');
        if (!sel) return;
        const placeholder = sel.options[0] ? sel.options[0].outerHTML : '<option value="">&#128269; 选择用例跳转...</option>';
        let html = placeholder;
        for (const item of titles) {
            const label = `${item.id} - ${item.title}`;
            html += `<option value="${item.index}">${escapeHtml(label)}</option>`;
        }
        sel.innerHTML = html;
    } catch (err) {
        console.error('加载用例列表失败:', err);
    }
}

function syncCaseJumpSelect(index) {
    const sel = document.getElementById('caseJumpSelect');
    if (!sel) return;
    sel.value = String(index);
}

// ============================================================
// 进度条
// ============================================================
async function refreshStatuses() {
    try {
        const resp = await fetch('/api/all-status');
        allStatuses = await resp.json();
        renderProgress();
    } catch (err) {
        console.error('获取状态失败:', err);
    }
}

function renderProgress() {
    const total = allStatuses.length;
    if (total === 0) return;
    const counts = {'通过':0,'失败':0,'阻塞':0,'跳过':0,'':0};
    for (const s of allStatuses) {
        if (counts.hasOwnProperty(s)) counts[s]++;
        else counts['']++;
    }
    const executed = total - counts[''];
    const pct = Math.round((executed / total) * 100);

    document.getElementById('progressText').textContent = `已执行 ${executed} / ${total} 条`;
    document.getElementById('progressPercent').textContent = pct + '%';
    document.getElementById('progressFill').style.width = pct + '%';
    document.getElementById('progressFill').className = 'progress-fill' + (pct === 100 ? ' done' : '');
    document.getElementById('countPass').textContent = counts['通过'];
    document.getElementById('countFail').textContent = counts['失败'];
    document.getElementById('countBlock').textContent = counts['阻塞'];
    document.getElementById('countSkip').textContent = counts['跳过'];
    document.getElementById('countNone').textContent = counts[''];
}

// ============================================================
// 搜索筛选
// ============================================================
let searchDebounceTimer = null;

function doSearch() {
    clearTimeout(searchDebounceTimer);
    searchDebounceTimer = setTimeout(_doSearch, 200);
}

async function _doSearch() {
    const keyword = document.getElementById('searchKeyword').value.trim();
    const resultFilter = document.getElementById('filterResult').value;
    const priorityFilter = document.getElementById('filterPriority').value;
    const moduleFilter = document.getElementById('filterModule').value;

    searchLastParams = {keyword, resultFilter, priorityFilter, moduleFilter, page:searchPage};

    try {
        const resp = await fetch('/api/search', {
            method:'POST',
            headers:{'Content-Type':'application/json'},
            body:JSON.stringify({
                keyword, result_filter:resultFilter,
                priority_filter:priorityFilter, module_filter:moduleFilter,
                page:searchPage, page_size:searchPageSize,
            }),
        });
        const data = await resp.json();
        renderSearchResults(data);
    } catch (err) {
        console.error('搜索失败:', err);
    }
}

function renderSearchResults(data) {
    const container = document.getElementById('searchContainer');
    const {results, total_matched, page, total_pages} = data;

    document.getElementById('searchInfo').textContent =
        total_matched > 0 ? `找到 ${total_matched} 条匹配用例` : '';

    if (results.length === 0) {
        container.innerHTML = `
            <div class="state-message">
                <div class="icon">&#128269;</div>
                <h2>无匹配结果</h2>
                <p>尝试更换关键词或放宽筛选条件</p>
            </div>`;
        document.getElementById('searchPagination').style.display = 'none';
        return;
    }

    // 高亮关键词
    const kw = (searchLastParams && searchLastParams.keyword) ? searchLastParams.keyword.toLowerCase() : '';

    let html = '<div class="search-result-list">';
    for (const r of results) {
        const badge = r.result
            ? `<span class="result-badge result-badge-${resultCss(r.result)}">${escapeHtml(r.result)}</span>`
            : '';
        const snippet = r.snippet ? highlightText(r.snippet, kw) : '';
        const titleDisplay = kw ? highlightText(r.name || r.title, kw) : escapeHtml(r.name || r.title);
        const idDisplay = r.id || '#' + (r.index + 1);

        html += `
            <div class="search-result-item" onclick="jumpToExecute(${r.index})" title="点击跳转到该用例">
                <div class="sr-left">
                    <span class="sr-id">${escapeHtml(idDisplay)}</span>
                    <div class="sr-title">${titleDisplay}</div>
                    <div class="sr-meta">
                        ${r.module ? escapeHtml(r.module) + ' &middot; ' : ''}
                        ${r.priority ? escapeHtml(r.priority) : ''}
                    </div>
                    ${snippet ? `<div class="sr-snippet">${snippet}</div>` : ''}
                </div>
                <div class="sr-right">
                    ${badge}
                    <span style="font-size:12px;color:var(--text-secondary);">#${r.index + 1}</span>
                </div>
            </div>`;
    }
    html += '</div>';
    container.innerHTML = html;

    // 分页
    const pagDiv = document.getElementById('searchPagination');
    if (total_pages > 1) {
        pagDiv.style.display = 'flex';
        pagDiv.innerHTML = `
            <button class="btn btn-outline btn-sm" onclick="searchGoPage(0)" ${page === 0 ? 'disabled' : ''}>首页</button>
            <button class="btn btn-outline btn-sm" onclick="searchGoPage(${page - 1})" ${page === 0 ? 'disabled' : ''}>上一页</button>
            <span class="page-info">第 ${page + 1} / ${total_pages} 页（共 ${total_matched} 条）</span>
            <button class="btn btn-outline btn-sm" onclick="searchGoPage(${page + 1})" ${page >= total_pages - 1 ? 'disabled' : ''}>下一页</button>
            <button class="btn btn-outline btn-sm" onclick="searchGoPage(${total_pages - 1})" ${page >= total_pages - 1 ? 'disabled' : ''}>末页</button>
        `;
    } else {
        pagDiv.style.display = 'none';
    }
}

function searchGoPage(p) {
    searchPage = p;
    if (searchLastParams) searchLastParams.page = p;
    _doSearch();
    window.scrollTo({top:0, behavior:'smooth'});
}

function clearSearch() {
    document.getElementById('searchKeyword').value = '';
    document.getElementById('filterResult').value = '';
    document.getElementById('filterPriority').value = '';
    document.getElementById('filterModule').value = '';
    searchPage = 0;
    doSearch();
}

function highlightText(text, keyword) {
    if (!keyword) return escapeHtml(text);
    const escaped = escapeHtml(text);
    const escapedKw = escapeHtml(keyword);
    const regex = new RegExp('(' + escapedKw.replace(/[.*+?^${}()|[\]\\]/g, '\\$&') + ')', 'gi');
    return escaped.replace(regex, '<em class="highlight">$1</em>');
}

function jumpToExecute(index) {
    switchView('execute').then ? switchView('execute').then(() => loadTestCase(index)) : (() => {
        _doSwitch('execute');
        loadTestCase(index);
    })();
}

async function loadFilterOptions() {
    try {
        const resp = await fetch('/api/filter-options');
        const data = await resp.json();
        // 填充优先级下拉
        const priSel = document.getElementById('filterPriority');
        priSel.innerHTML = '<option value="">全部优先级</option>';
        for (const p of data.priorities) {
            priSel.innerHTML += `<option value="${escapeHtml(p)}">${escapeHtml(p)}</option>`;
        }
        // 填充模块下拉
        const modSel = document.getElementById('filterModule');
        modSel.innerHTML = '<option value="">全部模块</option>';
        for (const m of data.modules) {
            modSel.innerHTML += `<option value="${escapeHtml(m)}">${escapeHtml(m)}</option>`;
        }
    } catch (err) {
        console.error('加载筛选选项失败:', err);
    }
}

// ============================================================
// 汇总统计
// ============================================================
async function loadSummary() {
    const container = document.getElementById('summaryContainer');
    container.innerHTML = '<div class="state-message"><div class="icon">⏳</div><h2>加载中...</h2></div>';

    try {
        const resp = await fetch('/api/summary');
        const data = await resp.json();
        renderSummary(data);
    } catch (err) {
        container.innerHTML = '<div class="state-message"><div class="icon">❌</div><h2>加载失败</h2></div>';
    }
}

function renderSummary(data) {
    const container = document.getElementById('summaryContainer');
    const {total, counts, by_priority, by_module, execution_rate, pass_rate} = data;

    // 环形进度条 SVG
    const ringSVG = (pct, color) => {
        const r = 26, circ = 2 * Math.PI * r;
        const offset = circ * (1 - pct / 100);
        return `<svg width="64" height="64" viewBox="0 0 64 64">
            <circle class="bg" cx="32" cy="32" r="${r}"/>
            <circle class="fg" cx="32" cy="32" r="${r}"
                stroke="${color}" stroke-dasharray="${circ}" stroke-dashoffset="${offset}"/>
        </svg>`;
    };

    let html = '';

    // 统计卡片
    html += `<div class="summary-grid">
        <div class="summary-stat stat-total"><div class="stat-number">${total}</div><div class="stat-label">用例总数</div></div>
        <div class="summary-stat stat-pass"><div class="stat-number">${counts['通过']}</div><div class="stat-label">通过</div></div>
        <div class="summary-stat stat-fail"><div class="stat-number">${counts['失败']}</div><div class="stat-label">失败</div></div>
        <div class="summary-stat stat-block"><div class="stat-number">${counts['阻塞']}</div><div class="stat-label">阻塞</div></div>
        <div class="summary-stat stat-skip"><div class="stat-number">${counts['跳过']}</div><div class="stat-label">跳过</div></div>
    </div>`;

    // 指标行
    html += `<div class="summary-metrics">
        <div class="metric-card">
            <div class="metric-ring">
                ${ringSVG(execution_rate, '#4f46e5')}
                <div class="pct" style="color:#4f46e5;">${execution_rate}%</div>
            </div>
            <div class="metric-text">
                <h4>执行进度</h4>
                <p>已执行 ${total - counts['未执行']} / ${total} 条</p>
            </div>
        </div>
        <div class="metric-card">
            <div class="metric-ring">
                ${ringSVG(pass_rate, '#16a34a')}
                <div class="pct" style="color:#16a34a;">${pass_rate}%</div>
            </div>
            <div class="metric-text">
                <h4>通过率</h4>
                <p>已执行用例中的通过比例</p>
            </div>
        </div>
    </div>`;

    // 按优先级汇总
    const priKeys = Object.keys(by_priority).sort();
    if (priKeys.length > 0) {
        html += `<div class="summary-table-wrap"><h3>⚡ 按优先级分布</h3>
        <table class="summary-table">
            <tr><th>优先级</th><th>用例数</th><th style="color:var(--pass)">通过</th><th style="color:var(--fail)">失败</th><th style="color:var(--block)">阻塞</th><th>跳过</th><th>未执行</th><th>进度</th></tr>`;
        for (const p of priKeys) {
            const d = by_priority[p];
            const totalP = d['通过'] + d['失败'] + d['阻塞'] + d['跳过'] + d['未执行'];
            const execP = totalP - d['未执行'];
            const pctP = totalP > 0 ? Math.round((execP / totalP) * 100) : 0;
            html += `<tr>
                <td><strong>${escapeHtml(p)}</strong></td>
                <td>${totalP}</td>
                <td class="cell-pass">${d['通过']}</td>
                <td class="cell-fail">${d['失败']}</td>
                <td class="cell-block">${d['阻塞']}</td>
                <td>${d['跳过']}</td>
                <td>${d['未执行']}</td>
                <td class="bar-cell">
                    <div class="mini-bar"><div class="mini-bar-fill pass" style="width:${pctP}%"></div></div>
                    <span style="font-size:11px;">${pctP}%</span>
                </td>
            </tr>`;
        }
        html += `</table></div>`;
    }

    // 按模块汇总
    const modKeys = Object.keys(by_module).sort();
    if (modKeys.length > 0) {
        html += `<div class="summary-table-wrap"><h3>&#128193; 按模块分布</h3>
        <table class="summary-table">
            <tr><th>模块</th><th>用例数</th><th style="color:var(--pass)">通过</th><th style="color:var(--fail)">失败</th><th style="color:var(--block)">阻塞</th><th>跳过</th><th>未执行</th><th>进度</th></tr>`;
        for (const m of modKeys) {
            const d = by_module[m];
            const totalM = d['通过'] + d['失败'] + d['阻塞'] + d['跳过'] + d['未执行'];
            const execM = totalM - d['未执行'];
            const pctM = totalM > 0 ? Math.round((execM / totalM) * 100) : 0;
            html += `<tr>
                <td><strong>${escapeHtml(m)}</strong></td>
                <td>${totalM}</td>
                <td class="cell-pass">${d['通过']}</td>
                <td class="cell-fail">${d['失败']}</td>
                <td class="cell-block">${d['阻塞']}</td>
                <td>${d['跳过']}</td>
                <td>${d['未执行']}</td>
                <td class="bar-cell">
                    <div class="mini-bar"><div class="mini-bar-fill pass" style="width:${pctM}%"></div></div>
                    <span style="font-size:11px;">${pctM}%</span>
                </td>
            </tr>`;
        }
        html += `</table></div>`;
    }

    container.innerHTML = html;
}

// ============================================================
// 未保存提示
// ============================================================
function showUnsavedDialog() {
    return new Promise((resolve) => {
        const overlay = document.createElement('div');
        overlay.className = 'modal-overlay';
        overlay.innerHTML = `
            <div class="modal-box">
                <h3>⚠️ 有未保存的更改</h3>
                <p>当前用例的结果或备注尚未保存，<br>切换后将丢失这些更改。是否先保存？</p>
                <div class="modal-actions">
                    <button class="btn-modal-ghost" id="modalDiscard">不保存，直接切换</button>
                    <button class="btn-modal-primary" id="modalSave">先保存</button>
                </div>
            </div>`;
        document.body.appendChild(overlay);

        overlay.querySelector('#modalSave').onclick = async () => {
            document.body.removeChild(overlay);
            await saveResult();
            resolve(true);
        };
        overlay.querySelector('#modalDiscard').onclick = () => {
            document.body.removeChild(overlay);
            dirty = false;
            resolve(true);
        };
        overlay.addEventListener('click', (e) => {
            if (e.target === overlay) { document.body.removeChild(overlay); resolve(false); }
        });
    });
}

// ============================================================
// Toast
// ============================================================
function showToast(msg, type) {
    const container = document.getElementById('toastContainer');
    const toast = document.createElement('div');
    toast.className = 'toast toast-' + type;
    toast.textContent = msg;
    container.appendChild(toast);
    setTimeout(() => { if (toast.parentNode) toast.parentNode.removeChild(toast); }, 2300);
}

function escapeHtml(str) {
    const div = document.createElement('div');
    div.appendChild(document.createTextNode(str));
    return div.innerHTML;
}

// ============================================================
// 刷新数据（重新加载 Excel）
// ============================================================
async function reloadData() {
    const btn = document.getElementById('btnReload');
    btn.disabled = true;
    btn.textContent = '⏳ 刷新中...';

    try {
        const resp = await fetch('/api/reload');
        const data = await resp.json();

        if (data.success) {
            totalCount = data.total;
            document.getElementById('topbarFilename').textContent =
                '📂 ' + (data.filename || '已加载');
            document.getElementById('progressWrap').style.display = 'block';
            document.getElementById('navBar').style.display = 'flex';
            document.getElementById('navHint').style.display = 'block';

            // 重置到第一条
            currentIndex = 0;
            selectedResult = '';
            dirty = false;

            await refreshStatuses();
            await loadTestCase(0);

            // 如果当前在搜索或汇总页，也刷新
            if (currentView === 'search') {
                await loadFilterOptions();
                searchPage = 0;
                await _doSearch();
            } else if (currentView === 'summary') {
                await loadSummary();
            }

            showToast('✅ 已刷新！共 ' + data.total + ' 条用例', 'success');
        } else {
            showToast('❌ 刷新失败: ' + (data.error || '未知错误'), 'error');
        }
    } catch (err) {
        showToast('❌ 刷新失败: ' + err.message, 'error');
    } finally {
        btn.disabled = false;
        btn.textContent = '↻ 刷新';
    }
}

// ============================================================
// 键盘快捷键
// ============================================================
document.addEventListener('keydown', function(e) {
    if (currentView !== 'execute') return;

    const tag = document.activeElement.tagName;
    const isInput = tag === 'INPUT' || tag === 'TEXTAREA' || tag === 'SELECT';

    if (!isInput) {
        if (e.key === 'ArrowLeft') { e.preventDefault(); goTo(-1); return; }
        if (e.key === 'ArrowRight') { e.preventDefault(); goTo(1); return; }
    }

    if ((e.ctrlKey || e.metaKey) && e.key === 's') {
        e.preventDefault();
        if (currentView === 'execute') saveResult();
    }
});

// ============================================================
// 问题列表
// ============================================================
async function loadIssues() {
    const container = document.getElementById('issuesContainer');
    container.innerHTML = '<div class="state-message"><div class="icon">⏳</div><h2>加载中...</h2></div>';

    try {
        const resp = await fetch('/api/all-status');
        const statuses = await resp.json();
        const total = statuses.length;

        // 统计
        let passCount = 0, failCount = 0, blockCount = 0, skipCount = 0, noneCount = 0;
        const issueIndices = [];
        for (let i = 0; i < statuses.length; i++) {
            const s = statuses[i];
            if (s === '通过') passCount++;
            else if (s === '失败') { failCount++; issueIndices.push({index: i, result: s}); }
            else if (s === '阻塞') { blockCount++; issueIndices.push({index: i, result: s}); }
            else if (s === '跳过') skipCount++;
            else noneCount++;
        }

        // 判断逻辑:
        // 1. 全部通过 (passCount == total) → 庆祝动画
        // 2. 全部执行完毕且没有失败/阻塞 (noneCount == 0 && failCount == 0 && blockCount == 0) → 庆祝动画
        // 3. 有失败或阻塞 → 展示问题列表
        // 4. 还没执行完毕，且没有失败阻塞 → 暂无问题
        const hasIssues = failCount > 0 || blockCount > 0;
        const executed = total - noneCount;
        const allPass = (passCount + skipCount === total && !hasIssues); // 通过+跳过=全部，没有失败阻塞

        if (allPass && noneCount === 0) {
            // 全部执行完毕且无失败阻塞 → 庆祝
            showCelebrate(container);
            return;
        }

        if (hasIssues) {
            // 有问题 → 展示列表
            let html = '<div style="margin-bottom:16px;">';
            html += '<h3 style="margin-bottom:8px;">&#9888; 共 ' + issueIndices.length + ' 条需要关注</h3>';
            html += '<p style="font-size:13px;color:var(--text-secondary);">以下用例未通过，修复后可在此页面进行回归测试</p>';
            html += '</div>';
            html += '<div class="issue-list">';

            for (const item of issueIndices) {
                try {
                    const tcResp = await fetch('/api/testcase/' + item.index);
                    const tc = await tcResp.json();
                    html += '<div class="issue-item" onclick="jumpToExecute(' + item.index + ')">';
                    html += '<div class="issue-header">';
                    html += '<span class="issue-id">' + escapeHtml(tc.id || tc.col_0 || '#' + (item.index+1)) + '</span>';
                    html += '<span class="result-badge result-badge-' + resultCss(item.result) + '">' + escapeHtml(item.result) + '</span>';
                    html += '</div>';
                    html += '<div class="issue-title">' + escapeHtml(tc.name || tc.title || tc.col_2 || '(无标题)') + '</div>';
                    html += '<div class="issue-meta">';
                    if (tc.module) html += '<span>&#128193; ' + escapeHtml(tc.module) + '</span>';
                    if (tc.priority) html += '<span>&#9889; ' + escapeHtml(tc.priority) + '</span>';
                    if (tc._saved_bug_id) html += '<span>&#128030; ' + escapeHtml(tc._saved_bug_id) + '</span>';
                    if (tc._saved_issue_time) html += '<span>&#128339; ' + escapeHtml(tc._saved_issue_time) + '</span>';
                    html += '<span>#' + (item.index + 1) + '</span>';
                    html += '</div>';
                    html += '</div>';
                } catch (e) {
                    // skip individual errors
                }
            }
            html += '</div>';
            container.innerHTML = html;
        } else {
            // 还未执行完毕，且暂无失败阻塞 → 暂无问题
            container.innerHTML = '<div class="state-message">' +
                '<div class="icon">&#9989;</div>' +
                '<h2>暂无问题</h2>' +
                '<p>已执行 ' + executed + ' / ' + total + ' 条，目前没有失败或阻塞的用例。</p>' +
                '<p style="font-size:13px;color:var(--text-secondary);margin-top:4px;">完成全部测试后将自动判断结果</p>' +
                '</div>';
        }
    } catch (err) {
        container.innerHTML = '<div class="state-message"><div class="icon">&#10060;</div><h2>加载失败</h2></div>';
    }
}

function showCelebrate(container) {
    const colors = ['#22c55e','#4f46e5','#f59e0b','#ef4444','#3b82f6','#8b5cf6','#ec4899','#14b8a6'];
    let confettiHtml = '';
    for (let i = 0; i < 60; i++) {
        const color = colors[Math.floor(Math.random() * colors.length)];
        const dx = (Math.random() - 0.5) * 400;
        const dy = -(Math.random() * 300 + 50);
        const rot = (Math.random() - 0.5) * 720;
        const delay = Math.random() * 1.5;
        const size = 6 + Math.random() * 10;
        confettiHtml += '<div class="confetti" style="background:' + color +
            ';width:' + size + 'px;height:' + size + 'px' +
            ';left:50%;top:50%' +
            ';--dx:' + dx + 'px' +
            ';--dy:' + dy + 'px' +
            ';--rot:' + rot + 'deg' +
            ';animation-delay:' + delay + 's' +
            ';"></div>';
    }

    container.innerHTML = '<div class="celebrate-wrap">' +
        confettiHtml +
        '<div style="font-size:64px;z-index:1;">&#127881;</div>' +
        '<div class="celebrate-text">恭喜！所有测试全部通过！</div>' +
        '<div class="celebrate-sub">&#127942; 所有用例执行完毕且无失败、无阻塞，测试任务圆满完成</div>' +
        '</div>';
}

// ============================================================
// 启动
// ============================================================
init();
</script>

</body>
</html>
'''


# ============================================================
# 5. 启动入口
# ============================================================
def main():
    print("=" * 50)
    print("  🔬 测试用例记录表  v1.3.4")
    print("=" * 50)
    print()

    filepath = find_excel_file()
    if filepath is None:
        print("📋 未找到活跃文件，将在网页中引导上传")
        print("   首次使用请在浏览器中上传测试用例 Excel")
        print()
    else:
        print(f"\U0001f4c2 找到活跃文件: {read_memory() or filepath.name}")
        try:
            loaded = read_all_sheets(filepath)
            _apply_loaded_data(filepath, loaded)

            print(f"\U0001f4ca 共读取 {len(STATE['testcases'])} 条测试用例")
            if STATE['mapping']:
                recognized = [k for k in STATE['mapping'] if k != '_headers']
                if recognized:
                    print(f"\U0001f50d 识别到的字段: {', '.join(recognized)}")
            print()
        except Exception as e:
            print(f"❌ 读取 Excel 失败: {e}")
            print("   将在网页中引导重新上传")
            print()

    print(f"\U0001f310 正在启动本地服务 (http://127.0.0.1:{PORT}) ...")
    print()

    def open_browser():
        import time
        time.sleep(0.8)
        webbrowser.open(f'http://127.0.0.1:{PORT}')

    threading.Thread(target=open_browser, daemon=True).start()

    import logging
    log = logging.getLogger('werkzeug')
    log.setLevel(logging.WARNING)

    print(f"✅ 服务已启动！浏览器正在打开...")
    print(f"   如果浏览器未自动打开，请手动访问: http://127.0.0.1:{PORT}")
    print()
    print("   按 Ctrl+C 可停止服务")
    print("-" * 50)

    try:
        app.run(host='127.0.0.1', port=PORT, debug=False, use_reloader=False)
    except KeyboardInterrupt:
        print("\n\U0001f44b 已停止服务，再见！")


if __name__ == '__main__':
    main()
