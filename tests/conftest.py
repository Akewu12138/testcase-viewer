"""
共享测试夹具（fixtures）和配置。

在导入被测模块前需要做一些 monkey-patching，
因为 testcase_viewer.py 在模块顶层会：
1. 调用 _ensure_deps() → 检查依赖（测试环境下已安装，不会触发安装）
2. 创建 Flask app → 没问题
3. 但 webbrowser 和 threading 在 main() 中会用到，需要 mock 掉
"""

import sys
import os

# 阻止任何意外的浏览器打开
import webbrowser
webbrowser.open = lambda url: None

# 确保项目根目录在 Python 路径中
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

import pytest
import openpyxl
from openpyxl import Workbook


# ============================================================
# 导入被测模块
# ============================================================

from testcase_viewer import (
    find_excel_file,
    detect_columns,
    read_testcases,
    save_result,
    build_display_fields,
    app,
    STATE,
    TESTCASES_DIR,
    COLUMN_PATTERNS,
    RESULT_COL,
    REMARK_COL,
    TIME_COL,
)


# ============================================================
# 夹具
# ============================================================

@pytest.fixture
def app_client():
    """创建 Flask 测试客户端"""
    app.config['TESTING'] = True
    with app.test_client() as client:
        yield client


@pytest.fixture
def sample_workbook(tmp_path):
    """创建一个示例 Excel 文件，包含完整的测试用例数据，返回文件路径"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 表头
    headers = [
        "编号", "用例标题", "所属模块", "优先级", "前置条件",
        "测试步骤", "预期结果", "测试目的"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 数据行
    data = [
        ["TC001", "登录功能测试", "登录模块", "P0", "用户已注册",
         "1. 打开登录页\n2. 输入用户名密码\n3. 点击登录",
         "登录成功，跳转到首页", "验证正常登录流程"],
        ["TC002", "密码错误测试", "登录模块", "P1", "用户已注册",
         "1. 打开登录页\n2. 输入正确用户名和错误密码\n3. 点击登录",
         "提示密码错误，不跳转", "验证异常登录处理"],
        ["TC003", "空值提交测试", "登录模块", "P2", "",
         "1. 不输入任何内容\n2. 点击登录",
         "提示请输入用户名和密码", "验证空值处理"],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    filepath = tmp_path / "test_cases.xlsx"
    wb.save(str(filepath))
    return str(filepath)


@pytest.fixture
def sample_workbook_with_results(tmp_path):
    """创建已包含测试结果列的 Excel"""
    wb = Workbook()
    ws = wb.active

    headers = ["编号", "用例标题", "测试结果", "测试现象备注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    data = [
        ["TC001", "用例A", "通过", "一切正常"],
        ["TC002", "用例B", "失败", "报错了"],
        ["TC003", "用例C", "", ""],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    filepath = tmp_path / "test_cases_with_results.xlsx"
    wb.save(str(filepath))
    return str(filepath)


@pytest.fixture
def multi_sheet_workbook(tmp_path):
    """创建多 Sheet Excel：2 个用例 Sheet + 2 个非用例 Sheet。"""
    wb = Workbook()
    ws_note = wb.active
    ws_note.title = "说明"
    ws_note.append(["字段", "说明"])
    ws_note.append(["环境", "测试环境说明"])

    headers = ["编号", "用例标题", "所属模块", "优先级", "测试步骤", "预期结果"]

    ws_login = wb.create_sheet("登录用例")
    ws_login.append(headers)
    ws_login.append(["L001", "正常登录", "登录模块", "P0", "输入账号密码", "登录成功"])
    ws_login.append(["L002", "密码错误", "登录模块", "P1", "输入错误密码", "提示错误"])

    ws_pay = wb.create_sheet("支付用例")
    ws_pay.append(headers)
    ws_pay.append(["P001", "微信支付", "支付模块", "P0", "提交微信支付", "支付成功"])

    ws_summary = wb.create_sheet("统计")
    ws_summary.append(["项目", "数量"])
    ws_summary.append(["总数", 3])

    filepath = tmp_path / "multi_sheet_cases.xlsx"
    wb.save(str(filepath))
    return str(filepath)


@pytest.fixture
def sample_headers():
    """标准表头列表"""
    return ["编号", "用例标题", "所属模块", "优先级", "前置条件",
            "测试步骤", "预期结果", "测试目的"]


@pytest.fixture
def sample_mapping(sample_headers):
    """标准列映射"""
    return detect_columns(sample_headers)


@pytest.fixture
def sample_testcase(sample_workbook, sample_headers):
    """返回一条已解析的测试用例 dict"""
    testcases, mapping, headers, _, _ = read_testcases(sample_workbook)
    return testcases[0]


@pytest.fixture
def populated_state(sample_workbook):
    """预填充 STATE 以便测试 API（模拟真实加载后的完整 STATE）"""
    testcases, mapping, headers, sheet_name, header_row = read_testcases(sample_workbook)
    original = {
        'testcases': STATE['testcases'],
        'mapping': STATE['mapping'],
        'headers': STATE['headers'],
        'sheets': STATE.get('sheets', []),
        'testcase_sheets': STATE.get('testcase_sheets', []),
        'note_sheets': STATE.get('note_sheets', []),
        'filepath': STATE['filepath'],
        'filename': STATE['filename'],
        'sheet_name': STATE['sheet_name'],
        'header_row': STATE['header_row'],
    }
    STATE['testcases'] = testcases
    STATE['mapping'] = mapping
    STATE['headers'] = headers
    STATE['sheets'] = []
    STATE['testcase_sheets'] = []
    STATE['note_sheets'] = []
    STATE['filepath'] = sample_workbook
    STATE['filename'] = 'test_cases.xlsx'
    STATE['sheet_name'] = sheet_name
    STATE['header_row'] = header_row
    yield STATE
    # 还原
    STATE['testcases'] = original['testcases']
    STATE['mapping'] = original['mapping']
    STATE['headers'] = original['headers']
    STATE['sheets'] = original['sheets']
    STATE['testcase_sheets'] = original['testcase_sheets']
    STATE['note_sheets'] = original['note_sheets']
    STATE['filepath'] = original['filepath']
    STATE['filename'] = original['filename']
    STATE['sheet_name'] = original['sheet_name']
    STATE['header_row'] = original['header_row']


@pytest.fixture
def populated_multi_sheet_state(multi_sheet_workbook):
    """预填充多 Sheet STATE，用于验证 v2 Sheet 接口和跨 Sheet 保存。"""
    from testcase_viewer import read_all_sheets

    loaded = read_all_sheets(multi_sheet_workbook)
    original = {
        'testcases': STATE['testcases'],
        'mapping': STATE['mapping'],
        'headers': STATE['headers'],
        'sheets': STATE.get('sheets', []),
        'testcase_sheets': STATE.get('testcase_sheets', []),
        'note_sheets': STATE.get('note_sheets', []),
        'filepath': STATE['filepath'],
        'filename': STATE['filename'],
        'sheet_name': STATE['sheet_name'],
        'header_row': STATE['header_row'],
    }
    STATE['testcases'] = loaded['testcases']
    STATE['mapping'] = loaded['mapping']
    STATE['headers'] = loaded['headers']
    STATE['sheets'] = loaded['sheets']
    STATE['testcase_sheets'] = loaded['testcase_sheets']
    STATE['note_sheets'] = loaded['note_sheets']
    STATE['filepath'] = multi_sheet_workbook
    STATE['filename'] = 'multi_sheet_cases.xlsx'
    STATE['sheet_name'] = loaded['sheet_name']
    STATE['header_row'] = loaded['header_row']
    yield STATE
    STATE['testcases'] = original['testcases']
    STATE['mapping'] = original['mapping']
    STATE['headers'] = original['headers']
    STATE['sheets'] = original['sheets']
    STATE['testcase_sheets'] = original['testcase_sheets']
    STATE['note_sheets'] = original['note_sheets']
    STATE['filepath'] = original['filepath']
    STATE['filename'] = original['filename']
    STATE['sheet_name'] = original['sheet_name']
    STATE['header_row'] = original['header_row']
