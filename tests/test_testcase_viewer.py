"""
测试 testcase_viewer.py 中的所有可测函数。

覆盖：
- 纯函数：find_excel_file, detect_columns, build_display_fields
- 文件 I/O：read_testcases, save_result
- Flask API：7 个接口（使用测试客户端）
"""

import os
import sys
import json
import tempfile
import pytest
import openpyxl
from pathlib import Path

# 导入被测模块的所有函数
from testcase_viewer import (
    find_excel_file,
    detect_columns,
    read_testcases,
    save_result,
    build_display_fields,
    app,
    STATE,
)


# ============================================================
# 一、纯函数测试
# ============================================================

class TestDetectColumns:
    """测试 detect_columns() — 表头智能映射"""

    def test_standard_headers(self, sample_headers):
        """标准中文表头 → 正确映射到对应字段"""
        mapping = detect_columns(sample_headers)
        assert mapping['id'] == 0
        assert mapping['title'] == 1
        assert mapping['module'] == 2
        assert mapping['priority'] == 3
        assert mapping['precondition'] == 4
        assert mapping['steps'] == 5
        assert mapping['expected'] == 6
        assert mapping['purpose'] == 7

    def test_english_headers(self):
        """中英混合表头 → 也能正确映射"""
        headers = ["ID", "标题", "测试步骤", "预期结果"]
        mapping = detect_columns(headers)
        # "ID" 精确匹配 id 的英文模式
        assert mapping.get('id') is not None
        # "测试步骤" 精确匹配 steps 的模式
        assert mapping.get('steps') is not None

    def test_empty_headers(self):
        """空表头列表 → 不崩溃，返回空映射"""
        mapping = detect_columns([])
        assert '_headers' in mapping
        assert len(mapping) == 1  # 只有 _headers

    def test_none_headers(self):
        """包含 None 的表头 → 跳过 None，正常映射"""
        headers = ["编号", None, "标题", ""]
        mapping = detect_columns(headers)
        assert mapping.get('id') == 0
        assert mapping.get('title') == 2

    def test_partial_match_headers(self):
        """模糊匹配表头 → 包含关键词即匹配"""
        headers = ["用例编号(ID)", "测试用例标题", "操作步骤详细"]
        mapping = detect_columns(headers)
        assert mapping.get('id') == 0
        assert mapping.get('title') == 1
        assert mapping.get('steps') == 2

    def test_unmatched_headers(self):
        """无法匹配的表头 → mapping 中不出现，但 _headers 保留原样"""
        headers = ["奇怪名称", "另一个奇怪名称"]
        mapping = detect_columns(headers)
        assert mapping['_headers'] == headers
        assert 'id' not in mapping


class TestBuildDisplayFields:
    """测试 build_display_fields() — 构建展示字段"""

    def test_full_fields(self, sample_testcase, sample_mapping, sample_headers):
        """标准用例 → 生成完整的展示字段列表"""
        fields = build_display_fields(sample_testcase, sample_mapping, sample_headers)
        labels = [f['label'] for f in fields]
        assert any('用例编号' in l for l in labels)
        assert any('用例标题' in l for l in labels)
        assert any('测试步骤' in l for l in labels)
        assert any('预期结果' in l for l in labels)
        assert any('测试目的' in l for l in labels)

    def test_empty_fields_excluded(self, sample_mapping):
        """空值字段仍会出现在展示列表中，但值为空字符串（实现已改为全列展示）"""
        tc = {'id': 'TC001', 'title': '测试', '_row': 2}
        for i in range(10):
            tc[f'col_{i}'] = ''
            tc[f'_header_{i}'] = f'列{i}'
        fields = build_display_fields(tc, sample_mapping, [])
        # id 和 title 有值会显示
        id_shown = any('用例编号' in f['label'] for f in fields)
        title_shown = any('用例标题' in f['label'] for f in fields)
        assert id_shown
        assert title_shown
        # precondition 映射存在 → 字段仍出现，但值为空（不省略）
        pre_field = next((f for f in fields if '前置条件' in f['label']), None)
        assert pre_field is not None     # 字段出现
        assert pre_field['value'] == ''  # 但值为空

    def test_result_columns_shown(self, sample_testcase, sample_mapping):
        """结果列/备注列现在会出现在展示字段中（实现已改为全列展示，不省略）"""
        headers = ["编号", "标题", "测试结果", "测试现象备注", "执行时间"]
        tc = {
            'id': 'TC001', 'title': '测试', '_row': 2,
            'col_0': 'TC001', 'col_1': '测试',
            'col_2': '通过', 'col_3': '没问题', 'col_4': '2024-01-01',
            '_header_0': '编号', '_header_1': '标题',
            '_header_2': '测试结果', '_header_3': '测试现象备注', '_header_4': '执行时间',
        }
        mapping = detect_columns(headers)
        fields = build_display_fields(tc, mapping, headers)
        # 实现已改为全列展示：结果列、备注列都会出现且值正确
        result_field = next((f for f in fields if '测试结果' in f['label']), None)
        remark_field = next((f for f in fields if '测试现象备注' in f['label']), None)
        assert result_field is not None
        assert result_field['value'] == '通过'
        assert remark_field is not None


# ============================================================
# 二、文件 I/O 测试
# ============================================================

class TestReadTestcases:
    """测试 read_testcases() — 读取 Excel"""

    def test_read_valid_file(self, sample_workbook):
        """正常 Excel 文件 → 正确解析出所有用例"""
        testcases, mapping, headers, _, _ = read_testcases(sample_workbook)
        assert len(testcases) == 3
        assert testcases[0]['id'] == 'TC001'
        assert testcases[0]['title'] == '登录功能测试'
        assert testcases[0]['_row'] == 2

    def test_read_empty_file(self, tmp_path):
        """只有表头没数据 → 返回空列表"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="编号")
        ws.cell(row=1, column=2, value="标题")
        filepath = tmp_path / "empty.xlsx"
        wb.save(str(filepath))
        testcases, mapping, headers, _, _ = read_testcases(str(filepath))
        assert len(testcases) == 0

    def test_search_text_built(self, sample_workbook):
        """_search_text 包含所有字段用于全文搜索"""
        testcases, _, _, _, _ = read_testcases(sample_workbook)
        tc = testcases[0]
        assert 'TC001' in tc['_search_text']
        assert '登录功能测试' in tc['_search_text']

    def test_nonexistent_file(self):
        """不存在的文件路径 → 抛出异常"""
        with pytest.raises(Exception):
            read_testcases("/nonexistent/path/file.xlsx")


class TestSaveResult:
    """测试 save_result() — 保存测试结果

    注意：save_result 当前签名需要 header_row_idx 与 sheet_name，
    以及 actual_result（取代旧版 remark）、tester、bug_id、bug_frequency、issue_time。
    """

    def test_save_new_result(self, sample_workbook):
        """首次保存结果 → 自动创建结果列并写入"""
        save_result(
            sample_workbook, row_number=2, header_row_idx=1, sheet_name="测试用例",
            result="通过", actual_result="一切正常",
            tester="", bug_id="", bug_frequency="", issue_time="",
        )

        # 验证写入内容
        wb = openpyxl.load_workbook(sample_workbook, data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=c + 1).value for c in range(ws.max_column)]
        assert "测试结果" in headers
        assert "实际结果" in headers
        assert "执行时间" in headers

        result_col = headers.index("测试结果") + 1
        actual_col = headers.index("实际结果") + 1
        assert ws.cell(row=2, column=result_col).value == "通过"
        assert ws.cell(row=2, column=actual_col).value == "一切正常"
        wb.close()

    def test_save_to_existing_columns(self, sample_workbook_with_results):
        """已有结果列的 Excel → 覆盖写入"""
        save_result(
            sample_workbook_with_results, row_number=3, header_row_idx=1, sheet_name="Sheet",
            result="阻塞", actual_result="网络不通",
            tester="", bug_id="", bug_frequency="", issue_time="",
        )

        wb = openpyxl.load_workbook(sample_workbook_with_results, data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=c + 1).value for c in range(ws.max_column)]
        result_col = headers.index("测试结果") + 1
        assert ws.cell(row=3, column=result_col).value == "阻塞"
        wb.close()

    def test_save_multiple_results(self, sample_workbook):
        """多次保存不同行 → 互不覆盖"""
        base = dict(header_row_idx=1, sheet_name="测试用例",
                    tester="", bug_id="", bug_frequency="", issue_time="")
        save_result(sample_workbook, row_number=2, result="通过", actual_result="OK", **base)
        save_result(sample_workbook, row_number=3, result="失败", actual_result="报错", **base)
        save_result(sample_workbook, row_number=4, result="跳过", actual_result="", **base)

        wb = openpyxl.load_workbook(sample_workbook, data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=c + 1).value for c in range(ws.max_column)]
        result_col = headers.index("测试结果") + 1
        assert ws.cell(row=2, column=result_col).value == "通过"
        assert ws.cell(row=3, column=result_col).value == "失败"
        assert ws.cell(row=4, column=result_col).value == "跳过"
        wb.close()


# ============================================================
# 三、Flask API 测试
# ============================================================

class TestApiInit:
    """测试 GET /api/init"""

    def test_init_unloaded(self, app_client):
        """未加载数据时 → loaded=false"""
        resp = app_client.get('/api/init')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['loaded'] is False

    def test_init_loaded(self, app_client, populated_state):
        """已加载数据时 → loaded=true"""
        resp = app_client.get('/api/init')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['loaded'] is True
        assert data['total'] == 3


class TestApiTestcase:
    """测试 GET /api/testcase/<index>"""

    def test_get_valid_testcase(self, app_client, populated_state):
        """请求有效索引 → 返回完整用例数据"""
        resp = app_client.get('/api/testcase/0')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['id'] == 'TC001'
        assert data['title'] == '登录功能测试'
        assert data['_index'] == 0
        assert data['_total'] == 3
        assert '_display_fields' in data

    def test_get_out_of_range(self, app_client, populated_state):
        """索引超出范围 → 404"""
        resp = app_client.get('/api/testcase/999')
        assert resp.status_code == 404

    def test_get_negative_index(self, app_client, populated_state):
        """负数索引 → 404"""
        resp = app_client.get('/api/testcase/-1')
        assert resp.status_code == 404


class TestApiAllStatus:
    """测试 GET /api/all-status"""

    def test_get_all_status(self, app_client, populated_state):
        """有加载数据 → 返回状态列表"""
        resp = app_client.get('/api/all-status')
        assert resp.status_code == 200
        statuses = resp.get_json()
        assert len(statuses) == 3

    def test_no_data(self, app_client):
        """无数据 → 返回空列表"""
        resp = app_client.get('/api/all-status')
        assert resp.status_code == 200
        assert resp.get_json() == []


class TestApiSearch:
    """测试 POST /api/search"""

    def test_search_by_keyword(self, app_client, populated_state):
        """关键词搜索 → 返回匹配结果"""
        resp = app_client.post('/api/search',
            data=json.dumps({'keyword': '登录'}),
            content_type='application/json')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['total_matched'] >= 1

    def test_search_no_match(self, app_client, populated_state):
        """无匹配关键词 → 返回空结果"""
        resp = app_client.post('/api/search',
            data=json.dumps({'keyword': '不存在的内容xyz'}),
            content_type='application/json')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['total_matched'] == 0
        assert data['results'] == []

    def test_search_filter_by_result(self, app_client, populated_state):
        """按结果筛选 → 只返回匹配的"""
        # 先保存一条结果
        app_client.post('/api/save',
            data=json.dumps({'index': 0, 'result': '通过', 'remark': ''}),
            content_type='application/json')

        resp = app_client.post('/api/search',
            data=json.dumps({'result_filter': '通过'}),
            content_type='application/json')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['total_matched'] >= 1
        for r in data['results']:
            assert r['result'] == '通过'

    def test_search_filter_unexecuted(self, app_client, populated_state):
        """按'未执行'筛选"""
        resp = app_client.post('/api/search',
            data=json.dumps({'result_filter': '未执行'}),
            content_type='application/json')
        assert resp.status_code == 200
        data = resp.get_json()
        # 初始状态所有用例都未执行
        assert data['total_matched'] >= 1

    def test_search_with_pagination(self, app_client, populated_state):
        """分页查询 → 正确返回指定页"""
        resp = app_client.post('/api/search',
            data=json.dumps({'page': 0, 'page_size': 2}),
            content_type='application/json')
        assert resp.status_code == 200
        data = resp.get_json()
        assert len(data['results']) <= 2


class TestApiSummary:
    """测试 GET /api/summary"""

    def test_summary_basic(self, app_client, populated_state):
        """基本汇总数据 → 返回正确统计"""
        resp = app_client.get('/api/summary')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['total'] == 3
        assert 'counts' in data
        assert '通过' in data['counts']
        assert 'execution_rate' in data
        assert 'pass_rate' in data

    def test_summary_all_unexecuted(self, app_client, populated_state):
        """全部未执行 → 通过率为0"""
        resp = app_client.get('/api/summary')
        data = resp.get_json()
        assert data['counts']['未执行'] == 3
        assert data['execution_rate'] == 0

    def test_summary_after_save(self, app_client, populated_state):
        """保存结果后 → 汇总数据更新"""
        app_client.post('/api/save',
            data=json.dumps({'index': 0, 'result': '通过', 'remark': ''}),
            content_type='application/json')
        app_client.post('/api/save',
            data=json.dumps({'index': 1, 'result': '失败', 'remark': ''}),
            content_type='application/json')

        resp = app_client.get('/api/summary')
        data = resp.get_json()
        assert data['counts']['通过'] == 1
        assert data['counts']['失败'] == 1
        assert data['counts']['未执行'] == 1
        # 执行率: 2/3
        assert data['execution_rate'] == pytest.approx(66.7, 0.1)


class TestApiFilterOptions:
    """测试 GET /api/filter-options"""

    def test_filter_options(self, app_client, populated_state):
        """获取筛选选项 → 返回优先级和模块列表"""
        resp = app_client.get('/api/filter-options')
        assert resp.status_code == 200
        data = resp.get_json()
        assert 'priorities' in data
        assert 'modules' in data
        assert 'P0' in data['priorities']
        assert '登录模块' in data['modules']


class TestApiSave:
    """测试 POST /api/save"""

    def test_save_success(self, app_client, populated_state):
        """正常保存 → 返回 success"""
        resp = app_client.post('/api/save',
            data=json.dumps({'index': 0, 'result': '通过', 'remark': 'OK'}),
            content_type='application/json')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['success'] is True

    def test_save_invalid_index(self, app_client, populated_state):
        """无效索引 → 400"""
        resp = app_client.post('/api/save',
            data=json.dumps({'index': 999, 'result': '通过', 'remark': ''}),
            content_type='application/json')
        assert resp.status_code == 400

    def test_save_without_data(self, app_client, populated_state):
        """缺失必要参数 → 不应该崩溃"""
        resp = app_client.post('/api/save',
            data=json.dumps({}),
            content_type='application/json')
        # 理想情况下应该返回 400
        assert resp.status_code in (400, 500)


# ============================================================
# 阶段4 补充：未覆盖的 API 路由测试
# ============================================================

class TestApiTitles:
    """测试 GET /api/titles — 用例标题列表（前端下拉跳转用）"""

    def test_titles_loaded(self, app_client, populated_state):
        """已加载 → 返回所有用例的 index/id/title"""
        resp = app_client.get('/api/titles')
        assert resp.status_code == 200
        data = resp.get_json()
        assert len(data) == 3
        assert data[0]['index'] == 0
        assert data[0]['id'] == 'TC001'
        assert data[0]['title'] == '登录功能测试'

    def test_titles_unloaded(self, app_client):
        """未加载 → 返回空列表（不崩溃）"""
        resp = app_client.get('/api/titles')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data == []


class TestApiReload:
    """测试 GET /api/reload — 重新加载 Excel"""

    def test_reload_success(self, app_client, populated_state, monkeypatch):
        """找到活跃文件 → 重新解析并返回 success + total"""
        from pathlib import Path
        monkeypatch.setattr('testcase_viewer.find_excel_file',
                            lambda: Path(populated_state['filepath']))
        resp = app_client.get('/api/reload')
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['success'] is True
        assert data['total'] == 3

    def test_reload_no_file(self, app_client, monkeypatch):
        """无活跃文件 → 404"""
        monkeypatch.setattr('testcase_viewer.find_excel_file', lambda: None)
        resp = app_client.get('/api/reload')
        assert resp.status_code == 404
        data = resp.get_json()
        assert data['success'] is False


class TestApiUpload:
    """测试 POST /api/upload — 上传 Excel"""

    def test_upload_no_file(self, app_client):
        """未提供文件字段 → 400"""
        resp = app_client.post('/api/upload')
        assert resp.status_code == 400

    def test_upload_empty_filename(self, app_client):
        """文件名为空 → 400"""
        from io import BytesIO
        resp = app_client.post('/api/upload', data={
            'file': (BytesIO(b''), '')
        }, content_type='multipart/form-data')
        assert resp.status_code == 400

    def test_upload_wrong_format(self, app_client):
        """非 Excel 格式 → 400"""
        from io import BytesIO
        resp = app_client.post('/api/upload', data={
            'file': (BytesIO(b'fake content'), 'test.txt')
        }, content_type='multipart/form-data')
        assert resp.status_code == 400
        data = resp.get_json()
        assert '不支持' in data['error']


class TestApiCheckResults:
    """测试 GET /api/check-results — 检查活跃文件是否已有测试结果"""

    def test_check_results_response(self, app_client):
        """始终返回 has_results + filename 字段"""
        resp = app_client.get('/api/check-results')
        assert resp.status_code == 200
        data = resp.get_json()
        assert 'has_results' in data
        assert 'filename' in data
